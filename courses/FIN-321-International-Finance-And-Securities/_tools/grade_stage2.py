"""FIN-321 Stage 2 grading scanner (FX Hedge Model).

Walks the Stage2 submissions directory, dedupes by student ID (keeps latest
timestamp), inspects each .xlsx, and produces a grading worksheet with rubric
signals and tentative auto-scores.

Rubric (6 points total):
  Structure & Clarity      2
  Accuracy                 2
  Sensitivity & Analysis   1
  Professionalism          1

Output goes to the same _grading/ folder so it stays inside ignore-term/.
"""
from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STAGE2_DIR = Path(
    r"C:\GitHub\shidler\courses\FIN-321-International-Finance-And-Securities"
    r"\ignore-term\2026-Spring\stage2"
)
OUTPUT_PATH = STAGE2_DIR / "_grading" / "stage2-grading-worksheet.xlsx"

# Folder name pattern: "<id>-<course> - <Name> - <Mon D, YYYY HHMM AM/PM>"
FOLDER_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)

# Two legitimate naming conventions observed:
#  A. Skeleton template (assignment spec):
SKELETON_NAMES = {
    "FC_AMT", "S0_in", "F0_in", "R_USD", "R_FC",
    "K_PUT", "K_CALL", "PREM_PUT", "PREM_CALL", "T_DAYS",
}
#  B. Descriptive lowercase alt (seen in many submissions):
ALT_NAME_TOKENS = {
    "call_price", "call_strike", "put_price", "put_strike",
    "forward_price", "current_spot_price", "future_spot_price",
    "contract_notional", "rate_us", "rate_uk", "recievable",
    "receivable", "payable", "scenario", "notional",
}

# Hedge-section keywords to detect on any sheet's cell text
HEDGE_KEYWORDS = {
    "Forward":      ["forward hedge", "forward rate", "forward contract", "locked-in", "locked in"],
    "MoneyMarket":  ["money market", "mm hedge", "borrow", "invest in usd"],
    "Put":          ["put option", "put hedge", "put premium", "strike (put)", "k_put", "put strike"],
    "Call":         ["call option", "call hedge", "call premium", "strike (call)", "k_call", "call strike"],
    "Sensitivity":  ["sensitivity", "±5", "+/-5", "scenario table", "s_t", "ending spot"],
}


@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime
    folder: Path
    xlsx: Path | None = None
    nonxlsx_note: str = ""


@dataclass
class Grade:
    student_id: str
    student_name: str
    submitted_at: datetime
    xlsx_name: str
    sheets: list[str] = field(default_factory=list)
    notes_tab: bool = False
    named_range_count: int = 0
    skeleton_name_hits: int = 0
    alt_name_hits: int = 0
    convention_rate: float = 0.0
    hedge_sections_found: list[str] = field(default_factory=list)
    chart_count: int = 0
    distinct_fill_colors: int = 0
    sensitivity_detected: bool = False
    auto_structure: int = 0        # /2
    auto_accuracy: int = 0         # /2
    auto_sensitivity: int = 0      # /1
    auto_professional: int = 0     # /1
    flags: list[str] = field(default_factory=list)
    error: str = ""


def parse_folder(folder: Path) -> Submission | None:
    m = FOLDER_RE.match(folder.name)
    if not m:
        return None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        return None
    return Submission(
        student_id=m.group("sid"),
        student_name=m.group("name").strip(),
        submitted_at=dt,
        folder=folder,
    )


def find_xlsx(folder: Path) -> tuple[Path | None, str]:
    """Return (best_xlsx, note) where note describes non-xlsx fallback if any."""
    xlsxs = [p for p in folder.iterdir()
             if p.is_file() and p.suffix.lower() == ".xlsx"]
    if xlsxs:
        xlsxs.sort(key=lambda p: len(p.name))
        return xlsxs[0], ""
    other = [p.name for p in folder.iterdir() if p.is_file()]
    if other:
        return None, "non-xlsx submission: " + ", ".join(other)
    return None, "empty folder"


def collect_submissions() -> list[Submission]:
    by_id: dict[str, Submission] = {}
    skipped = []
    for child in STAGE2_DIR.iterdir():
        if not child.is_dir() or child.name.startswith("_"):
            continue
        sub = parse_folder(child)
        if sub is None:
            skipped.append(child.name)
            continue
        existing = by_id.get(sub.student_id)
        if existing is None or sub.submitted_at > existing.submitted_at:
            by_id[sub.student_id] = sub
    if skipped:
        print(f"[warn] skipped {len(skipped)} unparsable folders:", skipped)
    subs = sorted(by_id.values(), key=lambda s: s.student_name.lower())
    for s in subs:
        s.xlsx, s.nonxlsx_note = find_xlsx(s.folder)
    return subs


def inspect(sub: Submission) -> Grade:
    g = Grade(
        student_id=sub.student_id,
        student_name=sub.student_name,
        submitted_at=sub.submitted_at,
        xlsx_name=sub.xlsx.name if sub.xlsx else sub.nonxlsx_note or "(none)",
    )
    if sub.xlsx is None:
        g.error = sub.nonxlsx_note or "no file"
        g.flags.append("NO_XLSX")
        return g

    try:
        wb = load_workbook(sub.xlsx, data_only=True, read_only=False)
    except Exception as e:
        g.error = f"open failed: {e}"
        g.flags.append("OPEN_FAILED")
        return g

    g.sheets = list(wb.sheetnames)
    lowered = [s.lower() for s in g.sheets]
    g.notes_tab = any("note" in s or "assumption" in s for s in lowered)

    names = [n for n in wb.defined_names]
    g.named_range_count = len(names)
    g.skeleton_name_hits = sum(1 for n in names if n in SKELETON_NAMES)
    lowered_names = [n.lower() for n in names]
    g.alt_name_hits = sum(
        1 for n in lowered_names
        if any(tok in n for tok in ALT_NAME_TOKENS)
    )
    best_convention_hits = max(g.skeleton_name_hits, g.alt_name_hits)
    g.convention_rate = (
        best_convention_hits / g.named_range_count if g.named_range_count else 0.0
    )

    text_blob = ""
    fills: set[str] = set()
    chart_count = 0
    for sname in g.sheets:
        ws = wb[sname]
        chart_count += len(getattr(ws, "_charts", []))
        max_r = min(ws.max_row or 1, 1200)
        max_c = min(ws.max_column or 1, 40)
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                if isinstance(cell.value, str):
                    text_blob += cell.value.lower() + " "
                try:
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.rgb:
                        rgb = str(fill.fgColor.rgb)
                        if rgb not in ("00000000", "FFFFFFFF"):
                            fills.add(rgb)
                except Exception:
                    pass
    g.chart_count = chart_count
    g.distinct_fill_colors = len(fills)

    found = []
    for section, keywords in HEDGE_KEYWORDS.items():
        if any(k in text_blob for k in keywords):
            found.append(section)
    g.hedge_sections_found = found
    g.sensitivity_detected = "Sensitivity" in found or chart_count > 0

    # Auto-scoring
    core_hedges = {"Forward", "MoneyMarket", "Put", "Call"}
    hedges_found = core_hedges & set(found)

    structure = 0
    if g.named_range_count >= 8 or g.skeleton_name_hits >= 6:
        structure += 1
    if g.distinct_fill_colors >= 2:
        structure += 1
    g.auto_structure = structure

    accuracy = 0
    if len(hedges_found) >= 3:
        accuracy += 1
    if len(hedges_found) == 4:
        accuracy += 1
    g.auto_accuracy = accuracy

    g.auto_sensitivity = 1 if (g.sensitivity_detected and chart_count > 0) else (
        1 if g.sensitivity_detected else 0
    )

    prof = 0
    if g.named_range_count >= 8 and g.convention_rate >= 0.50:
        prof += 1
    if g.notes_tab:
        # notes tab is a nice-to-have; keep professional flexible
        pass
    g.auto_professional = prof

    # Flags
    if g.named_range_count == 0:
        g.flags.append("NO_NAMED_RANGES")
    if g.skeleton_name_hits == 0 and g.alt_name_hits == 0:
        g.flags.append("NO_CONVENTION_MATCH")
    for req in ("Forward", "MoneyMarket", "Put", "Call"):
        if req not in found:
            g.flags.append(f"MISSING_{req.upper()}")
    if not g.sensitivity_detected:
        g.flags.append("NO_SENSITIVITY")
    if chart_count == 0:
        g.flags.append("NO_CHART")
    if not g.notes_tab:
        g.flags.append("NO_NOTES_TAB")

    return g


LETTER_GRADE_SCALE = [
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90),  ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80),  ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70),  ("D", 65, 67), ("F", 0, 65),
]


def _write_letter_grade_summary(sm, curved_col: int, generous_col: int,
                                first_row: int, last_row: int) -> None:
    """Append letter-grade distribution section to Summary sheet.

    Uses live COUNTIFS formulas against the Grading sheet Curved columns so
    the histogram updates whenever Final /6 is edited.
    """
    cv_l = get_column_letter(curved_col)
    gn_l = get_column_letter(generous_col)
    cv_rng = f"Grading!${cv_l}${first_row}:${cv_l}${last_row}"
    gn_rng = f"Grading!${gn_l}${first_row}:${gn_l}${last_row}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")

    sm.append([])
    headers = ["Letter", "Min %", "Min /6",
               "Count (Curved)", "Count (Curved 70%)",
               "Histogram (Curved)", "Histogram (Curved 70%)"]
    sm.append(headers)
    hr = sm.max_row
    for col_idx in range(1, len(headers) + 1):
        c = sm.cell(row=hr, column=col_idx)
        c.font = header_font
        c.fill = header_fill

    for letter, min_pct, max_pct in LETTER_GRADE_SCALE:
        min_pts = round(min_pct * 6 / 100, 4)
        r = sm.max_row + 1
        if letter == "F":
            max_pts = round(max_pct * 6 / 100, 4)
            cv_f = f'=COUNTIFS({cv_rng},">0",{cv_rng},"<"&{max_pts})'
            gn_f = f'=COUNTIFS({gn_rng},">0",{gn_rng},"<"&{max_pts})'
        elif max_pct is None:
            cv_f = f'=COUNTIF({cv_rng},">="&{min_pts})'
            gn_f = f'=COUNTIF({gn_rng},">="&{min_pts})'
        else:
            max_pts = round(max_pct * 6 / 100, 4)
            cv_f = f'=COUNTIFS({cv_rng},">="&{min_pts},{cv_rng},"<"&{max_pts})'
            gn_f = f'=COUNTIFS({gn_rng},">="&{min_pts},{gn_rng},"<"&{max_pts})'
        sm.cell(row=r, column=1, value=letter)
        sm.cell(row=r, column=2, value=min_pct)
        sm.cell(row=r, column=3, value=min_pts)
        sm.cell(row=r, column=4, value=cv_f)
        sm.cell(row=r, column=5, value=gn_f)
        sm.cell(row=r, column=6, value=f'=REPT("█",D{r})')
        sm.cell(row=r, column=7, value=f'=REPT("█",E{r})')

    r = sm.max_row + 1
    sm.cell(row=r, column=1, value="No submission")
    sm.cell(row=r, column=4, value=f'=COUNTIF({cv_rng},0)')
    sm.cell(row=r, column=5, value=f'=COUNTIF({gn_rng},0)')
    sm.cell(row=r, column=6, value=f'=REPT("█",D{r})')
    sm.cell(row=r, column=7, value=f'=REPT("█",E{r})')

    sm.column_dimensions["F"].width = 40
    sm.column_dimensions["G"].width = 40


def _write_curve_formulas(ws, final_col: int, rank_col: int,
                          quart_col: int, curved_col: int, generous_col: int,
                          first_row: int, last_row: int) -> None:
    """Write live Excel formulas for Rank, Quartile, Curved /6, Curved 70% /6.

    Ceiling-floor curve rounded up to nearest 0.25:
      Curved       = CEILING(MAX(Final, 50%-floor), 0.25)  -- bottom 3.00
      Curved 70%   = CEILING(MAX(Final, 70%-floor), 0.25)  -- bottom 4.20

    Never reduces a student's raw Final score; lifts bottom-rank submissions.
    Final = 0 → both columns = 0 (non-submission).
    """
    fl = get_column_letter(final_col)
    rl = get_column_letter(rank_col)
    rng = f"${fl}${first_row}:${fl}${last_row}"
    n_expr = f'COUNTIF({rng},">0")'

    curved_fill = PatternFill("solid", fgColor="E2EFDA")
    generous_fill = PatternFill("solid", fgColor="C6E0B4")
    for r in range(first_row, last_row + 1):
        final_ref = f"{fl}{r}"
        rank_ref = f"{rl}{r}"
        rank_f = (
            f'=IF({final_ref}=0,"",'
            f'RANK({final_ref},{rng},0)'
            f'+SUMPRODUCT(({rng}={final_ref})*(ROW({rng})<ROW())))'
        )
        ws.cell(row=r, column=rank_col, value=rank_f)
        quart_f = (
            f'=IF({final_ref}=0,"",'
            f'IF({rank_ref}<={n_expr}/4,1,'
            f'IF({rank_ref}<={n_expr}/2,2,'
            f'IF({rank_ref}<={n_expr}*3/4,3,4))))'
        )
        ws.cell(row=r, column=quart_col, value=quart_f)

        p = f'(({rank_ref}-1)/({n_expr}-1))'
        floor50 = (
            f'IF({p}<=0.25,6-{p}*4,'
            f'IF({p}<=0.5,5-({p}-0.25)*4,'
            f'IF({p}<=0.75,4-({p}-0.5)*2,'
            f'3.5-({p}-0.75)*2)))'
        )
        ws.cell(row=r, column=curved_col,
                value=f'=IF({final_ref}=0,0,CEILING(MAX({final_ref},{floor50}),0.25))'
                ).fill = curved_fill

        floor70 = (
            f'IF({p}<=0.25,6-{p}*2.4,'
            f'IF({p}<=0.5,5.4-({p}-0.25)*2.4,'
            f'IF({p}<=0.75,4.8-({p}-0.5)*1.2,'
            f'4.5-({p}-0.75)*1.2)))'
        )
        ws.cell(row=r, column=generous_col,
                value=f'=IF({final_ref}=0,0,CEILING(MAX({final_ref},{floor70}),0.25))'
                ).fill = generous_fill


def build_worksheet(grades: list[Grade]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"

    headers = [
        "Student ID", "Student Name", "Submitted", "File",
        "Sheets", "Notes Tab",
        "# Named Ranges", "Skeleton Hits (/10)", "Alt Hits", "Convention %",
        "Hedge Sections Found", "# Hedges",
        "# Charts", "Distinct Fill Colors", "Sensitivity Detected",
        "Auto Structure /2", "Auto Accuracy /2",
        "Auto Sensitivity /1", "Auto Professional /1",
        "Auto Total /6", "Final /6",
        "Rank", "Quartile", "Curved /6", "Curved 70% /6",
        "Flags", "Comments",
    ]
    ws.append(headers)

    FINAL_COL = headers.index("Final /6") + 1
    RANK_COL = headers.index("Rank") + 1
    QUART_COL = headers.index("Quartile") + 1
    CURVED_COL = headers.index("Curved /6") + 1
    GENEROUS_COL = headers.index("Curved 70% /6") + 1
    FLAGS_COL = headers.index("Flags") + 1

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")

    for g in grades:
        auto_total = (
            g.auto_structure + g.auto_accuracy
            + g.auto_sensitivity + g.auto_professional
        )
        row = [
            g.student_id,
            g.student_name,
            g.submitted_at.strftime("%Y-%m-%d %H:%M"),
            g.xlsx_name,
            ", ".join(g.sheets) if g.sheets else "",
            "Y" if g.notes_tab else "",
            g.named_range_count,
            g.skeleton_name_hits,
            g.alt_name_hits,
            round(g.convention_rate, 2),
            ", ".join(g.hedge_sections_found),
            len(g.hedge_sections_found),
            g.chart_count,
            g.distinct_fill_colors,
            "Y" if g.sensitivity_detected else "",
            g.auto_structure, g.auto_accuracy,
            g.auto_sensitivity, g.auto_professional,
            auto_total, auto_total,  # Final defaults to Auto Total for editing
            None, None, None, None,  # Rank, Quartile, Curved, Curved 70% — formulas below
            ", ".join(g.flags),
            g.error,
        ]
        ws.append(row)
        r = ws.max_row
        if g.error:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags:
            ws.cell(row=r, column=FLAGS_COL).fill = flag_fill

    _write_curve_formulas(ws, FINAL_COL, RANK_COL, QUART_COL, CURVED_COL,
                          GENEROUS_COL, 2, ws.max_row)

    widths = [11, 30, 17, 40, 38, 8,
              14, 18, 10, 12,
              36, 10, 10, 18, 18,
              16, 16, 18, 18, 14, 10,
              8, 10, 11, 13,
              34, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Summary sheet
    sm = wb.create_sheet("Summary")
    sm.append(["Metric", "Value"])
    sm.append(["Total unique students", len(grades)])
    sm.append(["Missing xlsx", sum(1 for g in grades if "NO_XLSX" in g.flags)])
    sm.append(["No named ranges", sum(1 for g in grades if "NO_NAMED_RANGES" in g.flags)])
    sm.append(["Missing Forward hedge", sum(1 for g in grades if "MISSING_FORWARD" in g.flags)])
    sm.append(["Missing MM hedge", sum(1 for g in grades if "MISSING_MONEYMARKET" in g.flags)])
    sm.append(["Missing Put hedge", sum(1 for g in grades if "MISSING_PUT" in g.flags)])
    sm.append(["Missing Call hedge", sum(1 for g in grades if "MISSING_CALL" in g.flags)])
    sm.append(["No sensitivity", sum(1 for g in grades if "NO_SENSITIVITY" in g.flags)])
    sm.append(["No chart", sum(1 for g in grades if "NO_CHART" in g.flags)])
    sm.append(["No notes tab", sum(1 for g in grades if "NO_NOTES_TAB" in g.flags)])
    if grades:
        avg = sum(
            g.auto_structure + g.auto_accuracy + g.auto_sensitivity + g.auto_professional
            for g in grades
        ) / len(grades)
        sm.append(["Average auto-score /6", round(avg, 2)])

    sm.append([])
    sm.append(["Curve policy", ""])
    sm.append(["Ranking basis", "Final /6 (updates live)"])
    sm.append(["Mode", "Ceiling-floor + round up: Curved = CEILING(MAX(Final, quartile floor), 0.25). Rounds up to nearest 0.25; never reduces raw score."])
    sm.append(["Tiebreak", "Sheet order (alphabetical by name)"])
    sm.append(["Q1 top 25% floor", "rank 1..N/4 → 6.00 → 5.00"])
    sm.append(["Q2 next 25% floor", "→ 5.00 → 4.00"])
    sm.append(["Q3 next 25% floor", "→ 4.00 → 3.50"])
    sm.append(["Q4 bottom 25% floor", "→ 3.50 → 3.00"])
    sm.append(["Final = 0", "Curved = 0 (non-submission)"])
    sm.append([])
    sm.append(["Curved 70% /6 (generous option)", "Same shape, floor = 4.20 (70% of 6)"])
    sm.append(["Q1 floor", "6.00 → 5.40"])
    sm.append(["Q2 floor", "5.40 → 4.80"])
    sm.append(["Q3 floor", "4.80 → 4.50"])
    sm.append(["Q4 floor", "4.50 → 4.20"])

    for col in (1, 2):
        sm.column_dimensions[get_column_letter(col)].width = 30
    for cell in sm[1]:
        cell.font = header_font
        cell.fill = header_fill

    _write_letter_grade_summary(sm, CURVED_COL, GENEROUS_COL, 2, ws.max_row)

    wb.save(OUTPUT_PATH)


def main() -> int:
    subs = collect_submissions()
    print(f"Found {len(subs)} unique students (deduped by ID).")
    grades: list[Grade] = []
    for s in subs:
        print(f"  grading {s.student_id} {s.student_name} ...", end=" ")
        g = inspect(s)
        grades.append(g)
        if g.error:
            print(f"ERROR: {g.error}")
        else:
            total = g.auto_structure + g.auto_accuracy + g.auto_sensitivity + g.auto_professional
            print(f"auto={total}/6 flags={','.join(g.flags) or '-'}")
    build_worksheet(grades)
    print(f"\nWrote {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
