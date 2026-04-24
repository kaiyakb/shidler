"""BUS-314 Stage 2 grading scanner.

Walks the Stage2 submissions directory, dedupes by student ID (keeping the
latest timestamp per ID), inspects each .xlsx, and produces a grading worksheet
with rubric signals and tentative auto-scores.

Output goes to the same _grading/ folder so it stays inside ignore-term/.
"""
from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STAGE2_DIR = Path(
    r"C:\GitHub\shidler\courses\BUS-314-International-Corporate-Finance"
    r"\ignore-term\2026 Spring\Stage2"
)
OUTPUT_PATH = STAGE2_DIR / "_grading" / "stage2-grading-worksheet.xlsx"

# Folder name pattern: "<id>-<course> - <Name> - <Mon D, YYYY HHMM AM/PM>"
FOLDER_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)

CONVENTION_PREFIXES = (
    "BAL_", "INC_", "CASH_", "RATIO_",
    "startYear_", "currentYear_", "avg_",
)
MARKET_NAMES = {
    "share_price", "shares_outstanding", "cost_capital",
    "tax_rate", "market_capitalization",
}

CATEGORY_KEYWORDS = {
    "Performance": ["market value added", "mva", "market-to-book",
                    "market to book", "economic value added", "eva"],
    "Profitability": ["roa", "roc", "roe", "return on assets",
                      "return on capital", "return on equity"],
    "Efficiency": ["asset turnover", "receivables turnover",
                   "collection period", "inventory turnover",
                   "days in inventory", "profit margin"],
    "Leverage": ["debt ratio", "debt-equity", "debt to equity",
                 "times interest", "cash coverage", "debt burden",
                 "leverage"],
    "Liquidity": ["nwc", "current ratio", "quick ratio", "cash ratio",
                  "working capital"],
    "Du Pont": ["du pont", "dupont"],
}


@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime
    folder: Path
    xlsx: Path | None = None


@dataclass
class Grade:
    student_id: str
    student_name: str
    submitted_at: datetime
    xlsx_name: str
    sheets: list[str] = field(default_factory=list)
    has_bs: bool = False
    has_is: bool = False
    has_cf: bool = False
    has_ratios: bool = False
    has_notes: bool = False
    named_range_count: int = 0
    convention_hits: int = 0
    convention_rate: float = 0.0
    ratio_named_count: int = 0
    ratio_categories_found: list[str] = field(default_factory=list)
    distinct_fill_colors_ratios: int = 0
    du_pont_roa: float | None = None
    direct_roa_start: float | None = None
    direct_roa_avg: float | None = None
    du_pont_roa_delta: float | None = None
    du_pont_roe: float | None = None
    direct_roe: float | None = None
    du_pont_roe_delta: float | None = None
    auto_structure: int = 0       # /2
    auto_accuracy: int = 0        # /2
    auto_named: int = 0           # /1
    auto_professional: int = 0    # /1 (flag for manual)
    flags: list[str] = field(default_factory=list)
    error: str = ""


def parse_folder(folder: Path) -> Submission | None:
    m = FOLDER_RE.match(folder.name)
    if not m:
        return None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        return None
    return Submission(
        student_id=m.group("sid"),
        student_name=m.group("name").strip(),
        submitted_at=dt,
        folder=folder,
    )


def find_xlsx(folder: Path) -> Path | None:
    candidates = [p for p in folder.iterdir()
                  if p.is_file() and p.suffix.lower() == ".xlsx"]
    if not candidates:
        return None
    candidates.sort(key=lambda p: len(p.name))
    return candidates[0]


def collect_submissions() -> list[Submission]:
    by_id: dict[str, Submission] = {}
    skipped = []
    for child in STAGE2_DIR.iterdir():
        if not child.is_dir() or child.name.startswith("_"):
            continue
        sub = parse_folder(child)
        if sub is None:
            skipped.append(child.name)
            continue
        existing = by_id.get(sub.student_id)
        if existing is None or sub.submitted_at > existing.submitted_at:
            by_id[sub.student_id] = sub
    if skipped:
        print(f"[warn] skipped {len(skipped)} unparsable folders:", skipped)
    subs = sorted(by_id.values(), key=lambda s: s.student_name.lower())
    for s in subs:
        s.xlsx = find_xlsx(s.folder)
    return subs


def _named_value(wb, name: str):
    try:
        dn = wb.defined_names[name]
    except KeyError:
        return None
    for sheet_name, coord in dn.destinations:
        coord = coord.replace("$", "")
        try:
            v = wb[sheet_name][coord].value
        except Exception:
            return None
        return v
    return None


def inspect(sub: Submission) -> Grade:
    g = Grade(
        student_id=sub.student_id,
        student_name=sub.student_name,
        submitted_at=sub.submitted_at,
        xlsx_name=sub.xlsx.name if sub.xlsx else "(none)",
    )
    if sub.xlsx is None:
        g.error = "no .xlsx in folder"
        g.flags.append("NO_FILE")
        return g

    try:
        wb = load_workbook(sub.xlsx, data_only=True, read_only=False)
    except Exception as e:
        g.error = f"open failed: {e}"
        g.flags.append("OPEN_FAILED")
        return g

    g.sheets = list(wb.sheetnames)
    lowered = [s.lower() for s in g.sheets]
    g.has_bs = any("balance" in s for s in lowered)
    g.has_is = any("income" in s for s in lowered)
    g.has_cf = any("cash flow" in s or "cashflow" in s for s in lowered)
    g.has_ratios = any("ratio" in s for s in lowered)
    g.has_notes = any("note" in s for s in lowered)

    names = [n for n in wb.defined_names]
    g.named_range_count = len(names)
    g.convention_hits = sum(
        1 for n in names
        if n.startswith(CONVENTION_PREFIXES) or n in MARKET_NAMES
    )
    g.convention_rate = (
        g.convention_hits / g.named_range_count if g.named_range_count else 0.0
    )
    g.ratio_named_count = sum(1 for n in names if n.startswith("RATIO_"))

    ratios_sheet = None
    for s in g.sheets:
        if "ratio" in s.lower():
            ratios_sheet = wb[s]
            break

    text_blob = ""
    fills: set[str] = set()
    if ratios_sheet is not None:
        for row in ratios_sheet.iter_rows(
            min_row=1, max_row=min(ratios_sheet.max_row or 1, 200),
            max_col=min(ratios_sheet.max_column or 1, 20),
        ):
            for cell in row:
                if isinstance(cell.value, str):
                    text_blob += cell.value.lower() + " "
                try:
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.rgb:
                        rgb = str(fill.fgColor.rgb)
                        if rgb not in ("00000000", "FFFFFFFF"):
                            fills.add(rgb)
                except Exception:
                    pass
    g.distinct_fill_colors_ratios = len(fills)

    found = []
    for cat, keywords in CATEGORY_KEYWORDS.items():
        if any(k in text_blob for k in keywords):
            found.append(cat)
    g.ratio_categories_found = found

    du_pont_roa = _named_value(wb, "RATIO_du_pont_roa")
    direct_roa_start = (
        _named_value(wb, "RATIO_roa_start_year")
        or _named_value(wb, "RATIO_roa_start")
        or _named_value(wb, "RATIO_roa")
    )
    direct_roa_avg = (
        _named_value(wb, "RATIO_roa_average")
        or _named_value(wb, "RATIO_roa_avg")
    )
    du_pont_roe = _named_value(wb, "RATIO_du_pont_roe")
    direct_roe = (
        _named_value(wb, "RATIO_roe_start_year")
        or _named_value(wb, "RATIO_roe_start")
        or _named_value(wb, "RATIO_roe")
    )

    def to_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    g.du_pont_roa = to_float(du_pont_roa)
    g.direct_roa_start = to_float(direct_roa_start)
    g.direct_roa_avg = to_float(direct_roa_avg)
    g.du_pont_roe = to_float(du_pont_roe)
    g.direct_roe = to_float(direct_roe)

    if g.du_pont_roa is not None:
        best = g.direct_roa_start if g.direct_roa_start is not None else g.direct_roa_avg
        if best is not None:
            g.du_pont_roa_delta = abs(g.du_pont_roa - best)
    if g.du_pont_roe is not None and g.direct_roe is not None:
        g.du_pont_roe_delta = abs(g.du_pont_roe - g.direct_roe)

    # Auto-scoring
    structure = 0
    if g.has_bs and g.has_is and g.has_cf and g.has_ratios:
        structure += 1
    elif sum([g.has_bs, g.has_is, g.has_cf, g.has_ratios]) >= 3:
        structure += 1  # lenient: partial credit if Ratios + 2 statements
    if g.distinct_fill_colors_ratios >= 2:
        structure += 1
    g.auto_structure = structure

    accuracy = 0
    if len(g.ratio_categories_found) >= 5:
        accuracy += 1
    if (
        (g.du_pont_roa_delta is not None and g.du_pont_roa_delta < 0.01)
        or (g.du_pont_roe_delta is not None and g.du_pont_roe_delta < 0.02)
    ):
        accuracy += 1
    g.auto_accuracy = accuracy

    g.auto_named = 1 if g.named_range_count >= 10 and g.convention_rate >= 0.70 else 0
    g.auto_professional = 1  # default full; flag in review column if issues

    if g.named_range_count == 0:
        g.flags.append("NO_NAMED_RANGES")
        g.auto_named = 0
        g.auto_professional = 0
    if not g.has_ratios:
        g.flags.append("NO_RATIOS_TAB")
    if not (g.has_bs and g.has_is and g.has_cf):
        g.flags.append("MISSING_STATEMENT")
    if g.ratio_named_count == 0 and g.has_ratios:
        g.flags.append("NO_RATIO_NAMES")
    if g.du_pont_roa_delta is not None and g.du_pont_roa_delta >= 0.01:
        g.flags.append("DU_PONT_ROA_MISMATCH")
    if g.du_pont_roe_delta is not None and g.du_pont_roe_delta >= 0.02:
        g.flags.append("DU_PONT_ROE_MISMATCH")
    if len(g.ratio_categories_found) < 5:
        g.flags.append("MISSING_CATEGORY")

    return g


LETTER_GRADE_SCALE = [
    # (letter, min_percent, max_percent_exclusive or None)
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90),  ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80),  ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70),  ("D", 65, 67), ("F", 0, 65),
]


def _write_letter_grade_summary(sm, curved_col: int, generous_col: int,
                                first_row: int, last_row: int) -> None:
    """Append letter-grade distribution section to Summary sheet.

    Uses live COUNTIFS formulas against the Grading sheet Curved columns so
    the histogram updates whenever Final /6 is edited.
    """
    cv_l = get_column_letter(curved_col)
    gn_l = get_column_letter(generous_col)
    cv_rng = f"Grading!${cv_l}${first_row}:${cv_l}${last_row}"
    gn_rng = f"Grading!${gn_l}${first_row}:${gn_l}${last_row}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")

    sm.append([])
    headers = ["Letter", "Min %", "Min /6",
               "Count (Curved)", "Count (Curved 70%)",
               "Histogram (Curved)", "Histogram (Curved 70%)"]
    sm.append(headers)
    hr = sm.max_row
    for col_idx in range(1, len(headers) + 1):
        c = sm.cell(row=hr, column=col_idx)
        c.font = header_font
        c.fill = header_fill

    for letter, min_pct, max_pct in LETTER_GRADE_SCALE:
        min_pts = round(min_pct * 6 / 100, 4)
        r = sm.max_row + 1
        if letter == "F":
            # Exclude non-submissions (Curved = 0); count only positive < threshold
            max_pts = round(max_pct * 6 / 100, 4)
            cv_f = f'=COUNTIFS({cv_rng},">0",{cv_rng},"<"&{max_pts})'
            gn_f = f'=COUNTIFS({gn_rng},">0",{gn_rng},"<"&{max_pts})'
        elif max_pct is None:
            cv_f = f'=COUNTIF({cv_rng},">="&{min_pts})'
            gn_f = f'=COUNTIF({gn_rng},">="&{min_pts})'
        else:
            max_pts = round(max_pct * 6 / 100, 4)
            cv_f = f'=COUNTIFS({cv_rng},">="&{min_pts},{cv_rng},"<"&{max_pts})'
            gn_f = f'=COUNTIFS({gn_rng},">="&{min_pts},{gn_rng},"<"&{max_pts})'
        sm.cell(row=r, column=1, value=letter)
        sm.cell(row=r, column=2, value=min_pct)
        sm.cell(row=r, column=3, value=min_pts)
        sm.cell(row=r, column=4, value=cv_f)
        sm.cell(row=r, column=5, value=gn_f)
        sm.cell(row=r, column=6, value=f'=REPT("█",D{r})')
        sm.cell(row=r, column=7, value=f'=REPT("█",E{r})')

    r = sm.max_row + 1
    sm.cell(row=r, column=1, value="No submission")
    sm.cell(row=r, column=4, value=f'=COUNTIF({cv_rng},0)')
    sm.cell(row=r, column=5, value=f'=COUNTIF({gn_rng},0)')
    sm.cell(row=r, column=6, value=f'=REPT("█",D{r})')
    sm.cell(row=r, column=7, value=f'=REPT("█",E{r})')

    # Widen histogram columns
    sm.column_dimensions["F"].width = 40
    sm.column_dimensions["G"].width = 40


def _write_curve_formulas(ws, final_col: int, rank_col: int,
                          quart_col: int, curved_col: int, generous_col: int,
                          first_row: int, last_row: int) -> None:
    """Write live Excel formulas for Rank, Quartile, Curved /6, Curved 70% /6.

    Ceiling-floor curve rounded up to nearest 0.25:
      Curved       = CEILING(MAX(Final, 50%-floor), 0.25)  -- bottom 3.00
      Curved 70%   = CEILING(MAX(Final, 70%-floor), 0.25)  -- bottom 4.20

    Never reduces a student's raw Final score; lifts bottom-rank submissions.
    Final = 0 → both columns = 0 (non-submission).
    """
    fl = get_column_letter(final_col)
    rl = get_column_letter(rank_col)
    rng = f"${fl}${first_row}:${fl}${last_row}"
    n_expr = f'COUNTIF({rng},">0")'

    curved_fill = PatternFill("solid", fgColor="E2EFDA")
    generous_fill = PatternFill("solid", fgColor="C6E0B4")
    for r in range(first_row, last_row + 1):
        final_ref = f"{fl}{r}"
        rank_ref = f"{rl}{r}"
        rank_f = (
            f'=IF({final_ref}=0,"",'
            f'RANK({final_ref},{rng},0)'
            f'+SUMPRODUCT(({rng}={final_ref})*(ROW({rng})<ROW())))'
        )
        ws.cell(row=r, column=rank_col, value=rank_f)
        quart_f = (
            f'=IF({final_ref}=0,"",'
            f'IF({rank_ref}<={n_expr}/4,1,'
            f'IF({rank_ref}<={n_expr}/2,2,'
            f'IF({rank_ref}<={n_expr}*3/4,3,4))))'
        )
        ws.cell(row=r, column=quart_col, value=quart_f)

        p = f'(({rank_ref}-1)/({n_expr}-1))'
        # Standard 50%-floor: 6 -> 5 -> 4 -> 3.5 -> 3
        floor50 = (
            f'IF({p}<=0.25,6-{p}*4,'
            f'IF({p}<=0.5,5-({p}-0.25)*4,'
            f'IF({p}<=0.75,4-({p}-0.5)*2,'
            f'3.5-({p}-0.75)*2)))'
        )
        ws.cell(row=r, column=curved_col,
                value=f'=IF({final_ref}=0,0,CEILING(MAX({final_ref},{floor50}),0.25))'
                ).fill = curved_fill

        # Generous 70%-floor: 6 -> 5.4 -> 4.8 -> 4.5 -> 4.2
        floor70 = (
            f'IF({p}<=0.25,6-{p}*2.4,'
            f'IF({p}<=0.5,5.4-({p}-0.25)*2.4,'
            f'IF({p}<=0.75,4.8-({p}-0.5)*1.2,'
            f'4.5-({p}-0.75)*1.2)))'
        )
        ws.cell(row=r, column=generous_col,
                value=f'=IF({final_ref}=0,0,CEILING(MAX({final_ref},{floor70}),0.25))'
                ).fill = generous_fill


def build_worksheet(grades: list[Grade]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"

    headers = [
        "Student ID", "Student Name", "Submitted", "File",
        "Sheets", "BS", "IS", "CF", "Ratios", "Notes",
        "# Named Ranges", "# Convention Hits", "Convention %",
        "# RATIO_*", "Ratios Tab Fill Colors",
        "Categories Found", "# Categories",
        "Du Pont ROA", "Direct ROA (start)", "Direct ROA (avg)", "|Δ| ROA",
        "Du Pont ROE", "Direct ROE", "|Δ| ROE",
        "Auto Structure /2", "Auto Accuracy /2",
        "Auto Named /1", "Auto Professional /1",
        "Auto Total /6", "Final /6",
        "Rank", "Quartile", "Curved /6", "Curved 70% /6",
        "Flags", "Comments",
    ]
    ws.append(headers)

    # Column positions (1-indexed)
    FINAL_COL = headers.index("Final /6") + 1
    RANK_COL = headers.index("Rank") + 1
    QUART_COL = headers.index("Quartile") + 1
    CURVED_COL = headers.index("Curved /6") + 1
    GENEROUS_COL = headers.index("Curved 70% /6") + 1
    FLAGS_COL = headers.index("Flags") + 1

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")

    for g in grades:
        auto_total = (
            g.auto_structure + g.auto_accuracy
            + g.auto_named + g.auto_professional
        )
        row = [
            g.student_id,
            g.student_name,
            g.submitted_at.strftime("%Y-%m-%d %H:%M"),
            g.xlsx_name,
            ", ".join(g.sheets) if g.sheets else "",
            "Y" if g.has_bs else "",
            "Y" if g.has_is else "",
            "Y" if g.has_cf else "",
            "Y" if g.has_ratios else "",
            "Y" if g.has_notes else "",
            g.named_range_count,
            g.convention_hits,
            round(g.convention_rate, 2),
            g.ratio_named_count,
            g.distinct_fill_colors_ratios,
            ", ".join(g.ratio_categories_found),
            len(g.ratio_categories_found),
            g.du_pont_roa, g.direct_roa_start, g.direct_roa_avg, g.du_pont_roa_delta,
            g.du_pont_roe, g.direct_roe, g.du_pont_roe_delta,
            g.auto_structure, g.auto_accuracy, g.auto_named, g.auto_professional,
            auto_total, auto_total,  # Final defaults to Auto Total for editing
            None, None, None, None,  # Rank, Quartile, Curved, Curved 70% — formulas below
            ", ".join(g.flags),
            g.error,
        ]
        ws.append(row)
        r = ws.max_row
        if g.error:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags:
            ws.cell(row=r, column=FLAGS_COL).fill = flag_fill

    _write_curve_formulas(ws, FINAL_COL, RANK_COL, QUART_COL, CURVED_COL,
                          GENEROUS_COL, 2, ws.max_row)

    widths = [11, 24, 17, 34, 38, 4, 4, 4, 7, 7,
              14, 16, 13, 10, 12, 30, 12,
              12, 18, 18, 10, 12, 12, 10,
              16, 16, 14, 18, 13, 10,
              8, 10, 11, 13,
              34, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Summary sheet
    sm = wb.create_sheet("Summary")
    sm.append(["Metric", "Value"])
    sm.append(["Total unique students", len(grades)])
    sm.append(["Missing file", sum(1 for g in grades if "NO_FILE" in g.flags)])
    sm.append(["No named ranges", sum(1 for g in grades if "NO_NAMED_RANGES" in g.flags)])
    sm.append(["Missing statement tab", sum(1 for g in grades if "MISSING_STATEMENT" in g.flags)])
    sm.append(["Missing ratio category", sum(1 for g in grades if "MISSING_CATEGORY" in g.flags)])
    sm.append(["Du Pont ROA mismatch", sum(1 for g in grades if "DU_PONT_ROA_MISMATCH" in g.flags)])
    sm.append(["Du Pont ROE mismatch", sum(1 for g in grades if "DU_PONT_ROE_MISMATCH" in g.flags)])
    if grades:
        avg = sum(
            g.auto_structure + g.auto_accuracy + g.auto_named + g.auto_professional
            for g in grades
        ) / len(grades)
        sm.append(["Average auto-score /6", round(avg, 2)])

    sm.append([])
    sm.append(["Curve policy", ""])
    sm.append(["Ranking basis", "Final /6 (updates live)"])
    sm.append(["Mode", "Ceiling-floor + round up: Curved = CEILING(MAX(Final, quartile floor), 0.25). Rounds up to nearest 0.25; never reduces raw score."])
    sm.append(["Tiebreak", "Sheet order (alphabetical by name)"])
    sm.append(["Q1 top 25% floor", "rank 1..N/4 → 6.00 → 5.00"])
    sm.append(["Q2 next 25% floor", "→ 5.00 → 4.00"])
    sm.append(["Q3 next 25% floor", "→ 4.00 → 3.50"])
    sm.append(["Q4 bottom 25% floor", "→ 3.50 → 3.00"])
    sm.append(["Final = 0", "Curved = 0 (non-submission)"])
    sm.append([])
    sm.append(["Curved 70% /6 (generous option)", "Same shape, floor = 4.20 (70% of 6)"])
    sm.append(["Q1 floor", "6.00 → 5.40"])
    sm.append(["Q2 floor", "5.40 → 4.80"])
    sm.append(["Q3 floor", "4.80 → 4.50"])
    sm.append(["Q4 floor", "4.50 → 4.20"])

    for col in (1, 2):
        sm.column_dimensions[get_column_letter(col)].width = 30
    for cell in sm[1]:
        cell.font = header_font
        cell.fill = header_fill

    _write_letter_grade_summary(sm, CURVED_COL, GENEROUS_COL, 2, ws.max_row)

    wb.save(OUTPUT_PATH)


def main() -> int:
    subs = collect_submissions()
    print(f"Found {len(subs)} unique students (deduped by ID).")
    grades: list[Grade] = []
    for s in subs:
        print(f"  grading {s.student_id} {s.student_name} ...", end=" ")
        g = inspect(s)
        grades.append(g)
        if g.error:
            print(f"ERROR: {g.error}")
        else:
            total = g.auto_structure + g.auto_accuracy + g.auto_named + g.auto_professional
            print(f"auto={total}/6 flags={','.join(g.flags) or '-'}")
    build_worksheet(grades)
    print(f"\nWrote {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
