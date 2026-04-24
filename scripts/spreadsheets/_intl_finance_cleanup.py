"""Full 4-phase cleanup for International Finance Spreadsheets.xlsx
  Phase 1 — error fixes + chapter section tabs
  Phase 2 — consistent coloring + Color Key + compact legends
  Phase 3 — Welcome tab + tab colors + hide data tabs
  Phase 4 — 5 resource tabs (Using This Workbook, Excel Formulas, Named
            Ranges, Data Analytics, Claude for Excel)

The workbook's final sheet order:
    Welcome / Color Key / Using This Workbook / Excel Formulas /
    Named Ranges / Data Analytics / Claude for Excel /
    [chapter section + member tabs, in original order]
"""
import re, shutil, zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.formula import ArrayFormula

SRC = Path(r'C:/GitHub/shidler/docs/spreadsheets/International Finance Spreadsheets.xlsx')

# ---- colors -----------------------------------------------------------------
UH_GREEN    = '024731'
UH_WHITE    = 'FFFFFF'
LIGHT_GREEN = 'E6EEEA'
INPUT_FILL  = 'FFF2CC'
FORMULA_FILL= 'F3F3F3'
FORMULA_FG  = '0000FF'
ARRAY_FG    = '9900FF'
LINK_FG     = '006100'
GRAY_TEXT   = '595959'
PROTECTED_FILLS = {'FFD9D9D9','FF9FC5E8','FFEFEFEF','FF024731','FFE6EEEA',
                   'FF000000'}

thin = Side(style='thin', color='BFBFBF')
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
bottom_rule = Border(bottom=Side(style='medium', color=UH_GREEN))

# Sheets to skip for the color-coding pass (pure data dumps)
SKIP_COLOR_TABS = {
    'GBTC','GLD','IEF','SPY','EWJ','EZU','EWA',
    'correl','data','ZM','UAL','Portfolio',
    'Correl ZM + UAL',
    'Equities','Equities wBTC','Asset Classes','Asset Classes wBTC',
}

# Sheets to hide from the tab strip (raw data)
HIDE_TABS = ['GBTC','GLD','IEF','SPY','EWJ','EZU','EWA','correl','data']

# Section tabs + their full titles + where their group ends
CHAPTER_MAP = [
    ('Chapter 6 International Parity',
     'Chapter 6 — International Parity',
     'Chapter 7 Options'),
    ('Chapter 7 Options',
     'Chapter 7 — Options & Derivatives',
     'Chapter 8 Transaction Hedging'),
    ('Chapter 8 Transaction Hedging',
     'Chapter 8 — Transaction Hedging (Receivables & Payables)',
     'Chapter 15  Intl Portffolio Man'),
    ('Chapter 15  Intl Portffolio Man',
     'Chapter 15 — International Portfolio Management',
     'Sharpe'),
    ('Sharpe',
     'Sharpe Ratios & Asset-Class Comparisons',
     None),
]
SECTION_TABS = {s[0] for s in CHAPTER_MAP}

# Tab colors by chapter
CHAPTER_TAB_COLORS = {
    'ch6':    '1F4E79',  # navy — parity
    'ch7':    '663398',  # purple — options
    'ch8':    '117A8B',  # teal — hedging
    'ch15':   '842029',  # deep red — portfolio mgmt
    'sharpe': '8B6F00',  # dark gold — Sharpe
}
CH_KEYS = ['ch6','ch7','ch8','ch15','sharpe']

wb = load_workbook(SRC, data_only=False)

# =============================================================================
# Phase 1 — error fixes + rebuild chapter section tabs
# =============================================================================

# (1a) _xludf.FORMULATEXT -> _xlfn.FORMULATEXT
xludf = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    if isinstance(ws, Chartsheet): continue
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and '_xludf.' in v:
                cell.value = v.replace('_xludf.', '_xlfn.')
                xludf += 1
print(f"[P1] _xludf replaced: {xludf}")

# (1b) #DIV/0! tail-row IFERROR wraps
div_fixes = [
    ('Correl ZM + UAL','I503'),('Correl ZM + UAL','L503'),
    ('Correl ZM + UAL','I504'),('Correl ZM + UAL','L504'),
    ('Equities','S120'),('Equities','S121'),('Equities','S122'),
    ('Equities wBTC','S120'),('Equities wBTC','S121'),('Equities wBTC','S122'),
    ('Asset Classes','S120'),('Asset Classes','S121'),('Asset Classes','S122'),
    ('Asset Classes wBTC','S120'),('Asset Classes wBTC','S121'),('Asset Classes wBTC','S122'),
]
iferr = 0
for sn, coord in div_fixes:
    c = wb[sn][coord]
    f = c.value
    if isinstance(f, str) and f.startswith('=') and 'IFERROR' not in f.upper():
        c.value = f'=IFERROR({f[1:]},"")'
        iferr += 1
print(f"[P1] IFERROR-wrapped: {iferr}")

# (1c) Rebuild chapter section tabs as styled dividers with TOC
full_order = list(wb.sheetnames)

def tabs_between(start_name, end_name):
    i0 = full_order.index(start_name) + 1
    return full_order[i0:] if end_name is None else full_order[i0:full_order.index(end_name)]

def clear_sheet(ws):
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

for section_tab, title, nxt in CHAPTER_MAP:
    ws = wb[section_tab]
    members = tabs_between(section_tab, nxt)
    clear_sheet(ws)

    # Title banner
    ws.merge_cells('B2:E2')
    t = ws['B2']; t.value = title
    t.font = Font(name='Open Sans', size=18, bold=True, color=UH_WHITE)
    t.fill = PatternFill('solid', start_color=UH_GREEN)
    t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 34

    ws.merge_cells('B3:E3')
    s = ws['B3']; s.value = 'Contents — click a tab name to jump to that sheet'
    s.font = Font(name='Open Sans', size=11, italic=True, color=GRAY_TEXT)
    s.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[3].height = 20

    # ToC header
    for i,h in enumerate(['#','Tab Name','Type'], start=2):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
        c.fill = PatternFill('solid', start_color=UH_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = box

    for idx, tab in enumerate(members, start=1):
        r = 5 + idx
        kind = 'Chart' if isinstance(wb[tab], Chartsheet) else 'Worksheet'
        n = ws.cell(row=r, column=2, value=idx)
        n.font = Font(name='Open Sans', size=11, color=GRAY_TEXT)
        n.alignment = Alignment(horizontal='center', vertical='center')
        n.border = box
        link = ws.cell(row=r, column=3, value=tab)
        safe = tab.replace("'", "''")
        link.hyperlink = f"#'{safe}'!A1"
        link.font = Font(name='Open Sans', size=11, color=UH_GREEN, underline='single')
        link.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        link.border = box
        tp = ws.cell(row=r, column=4, value=kind)
        tp.font = Font(name='Open Sans', size=11, color=GRAY_TEXT)
        tp.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        tp.border = box
        if idx % 2 == 0:
            for col in (2,3,4):
                ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 2
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.freeze_panes = 'B6'
    print(f"[P1] rebuilt {section_tab}: {len(members)} tabs")

# =============================================================================
# Phase 2 — coloring + Color Key + compact legends
# =============================================================================
def has_cross_sheet_ref(formula, current_sheet):
    refs = re.findall(r"'?([A-Za-z][^!'\"]*)'?!", formula)
    return any(s and s != current_sheet for s in refs)

def safe_fill_rgb(cell):
    try:
        v = cell.fill.fgColor.rgb
        if isinstance(v, str):
            return v.upper()
    except Exception:
        pass
    return None

# Apply coloring convention
stats = {'inputs':0,'formulas':0,'arrays':0,'links':0}
for sn in wb.sheetnames:
    if sn in SECTION_TABS or sn in SKIP_COLOR_TABS:
        continue
    ws = wb[sn]
    if isinstance(ws, Chartsheet): continue
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None: continue
            if cell.font.bold: continue
            existing = safe_fill_rgb(cell)
            if existing and existing in PROTECTED_FILLS:
                continue

            def new_font(color_rgb):
                return Font(
                    name='Open Sans', size=cell.font.size,
                    bold=cell.font.bold, italic=cell.font.italic,
                    underline=cell.font.underline, color=color_rgb,
                )

            is_array = isinstance(v, ArrayFormula)
            is_formula = isinstance(v, str) and v.startswith('=')

            if is_array:
                cell.fill = PatternFill('solid', start_color=FORMULA_FILL)
                cell.font = new_font(ARRAY_FG)
                stats['arrays'] += 1
            elif is_formula:
                if '_xlfn.FORMULATEXT(' in v:
                    cell.font = new_font(FORMULA_FG)
                    continue
                if has_cross_sheet_ref(v, sn):
                    cell.fill = PatternFill('solid', start_color=FORMULA_FILL)
                    cell.font = new_font(LINK_FG)
                    stats['links'] += 1
                else:
                    cell.fill = PatternFill('solid', start_color=FORMULA_FILL)
                    cell.font = new_font(FORMULA_FG)
                    stats['formulas'] += 1
            elif isinstance(v, (int, float)):
                cell.fill = PatternFill('solid', start_color=INPUT_FILL)
                cell.font = new_font(cell.font.color.rgb if (cell.font.color and isinstance(cell.font.color.rgb, str)) else None)
                stats['inputs'] += 1
            else:
                cell.font = new_font(cell.font.color.rgb if (cell.font.color and isinstance(cell.font.color.rgb, str)) else None)

print(f"[P2] coloring: {stats}")

# Color Key tab at position 0
if 'Color Key' in wb.sheetnames:
    del wb['Color Key']
key = wb.create_sheet('Color Key', 0)
key.sheet_view.showGridLines = False
key.sheet_view.zoomScale = 110
key.column_dimensions['A'].width = 2
key.column_dimensions['B'].width = 22
key.column_dimensions['C'].width = 26
key.column_dimensions['D'].width = 70
key.column_dimensions['E'].width = 2

key.merge_cells('B2:D2')
t = key['B2']; t.value = 'Color Key & Legend — International Finance Workbook'
t.font = Font(name='Open Sans', size=20, bold=True, color=UH_WHITE)
t.fill = PatternFill('solid', start_color=UH_GREEN)
t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
key.row_dimensions[2].height = 38

key.merge_cells('B3:D3')
s = key['B3']
s.value = ('This workbook accompanies FIN 321 International Finance & Securities. '
          'Tabs are grouped by chapter. Colors follow a consistent convention so '
          'you can tell at a glance which cells to edit and which to leave alone.')
s.font = Font(name='Open Sans', size=11, italic=True, color=GRAY_TEXT)
s.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
key.row_dimensions[3].height = 36

for col, text in [('B','Cell Type'),('C','Color'),('D','Meaning / When You Will See It')]:
    c = key[f'{col}5']; c.value = text
    c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
    c.fill = PatternFill('solid', start_color=UH_GREEN)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = box

legend_rows = [
    ('Input',         INPUT_FILL,   None,        '1.3415',
     'Hard-coded values — yellow fill means you can change these to explore scenarios.'),
    ('Formula',       FORMULA_FILL, FORMULA_FG,  '=invest*(1+rate_1y_US)',
     'Computed values. Blue text on gray fill. Do not edit.'),
    ('Array formula', FORMULA_FILL, ARRAY_FG,    '{=FREQUENCY(data,bins)}',
     'Dynamic/array formula. Purple text on gray fill.'),
    ('Linked cell',   FORMULA_FILL, LINK_FG,     "='ZM'!D10",
     'Formula that pulls a value from another worksheet. Green text on gray fill.'),
    ('Header / label',UH_GREEN,     UH_WHITE,    'CHAPTER 6',
     'Titles and section headings. UH green. Not calculated.'),
    ('Table header',  'D9D9D9',     '000000',    'Inputs  |  Outputs',
     'Column headings on data tables. Dark gray.'),
]
for i, (label, fill, fg, example, meaning) in enumerate(legend_rows):
    r = 6 + i; key.row_dimensions[r].height = 30
    a = key.cell(row=r, column=2, value=label)
    a.font = Font(name='Open Sans', size=11, bold=True)
    a.alignment = Alignment(horizontal='left', vertical='center', indent=1); a.border = box
    b = key.cell(row=r, column=3, value=example)
    b.fill = PatternFill('solid', start_color=fill)
    if fg:
        b.font = Font(name='Open Sans', size=11, color=fg, bold=(label=='Header / label'))
    else:
        b.font = Font(name='Open Sans', size=11)
    b.alignment = Alignment(horizontal='center', vertical='center'); b.border = box
    c = key.cell(row=r, column=4, value=meaning)
    c.font = Font(name='Open Sans', size=11, color='333333')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True); c.border = box

key.merge_cells('B13:D13')
tips = key['B13']; tips.value = 'Tip for students'
tips.font = Font(name='Open Sans', size=12, bold=True, color=UH_GREEN)
tips.alignment = Alignment(horizontal='left', vertical='center', indent=1)
tips.border = bottom_rule
key.merge_cells('B14:D16')
tb = key['B14']
tb.value = ('If a cell is yellow, you may change it to test a what-if. Everything '
            'else is either a label or a calculation that depends on other cells '
            '— editing those will break the model. Use the chapter tabs at the '
            'bottom of the workbook to jump to a topic.')
tb.font = Font(name='Open Sans', size=11, color='333333')
tb.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
key.freeze_panes = 'A5'
key.sheet_properties.tabColor = UH_GREEN
print("[P2] Color Key tab created")

# Compact legend on each chapter section tab
compact_rows = [
    ('Input',         INPUT_FILL,   None),
    ('Formula',       FORMULA_FILL, FORMULA_FG),
    ('Array formula', FORMULA_FILL, ARRAY_FG),
    ('Linked cell',   FORMULA_FILL, LINK_FG),
]
for sn in SECTION_TABS:
    ws = wb[sn]
    ws.merge_cells('G5:I5')
    h = ws['G5']; h.value = 'Color key'
    h.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
    h.fill = PatternFill('solid', start_color=UH_GREEN)
    h.alignment = Alignment(horizontal='left', vertical='center', indent=1); h.border = box
    for i,(label, fill, fg) in enumerate(compact_rows):
        r = 6 + i
        sw = ws.cell(row=r, column=7, value=''); sw.fill = PatternFill('solid', start_color=fill); sw.border = box
        nm = ws.cell(row=r, column=8, value=label); nm.font = Font(name='Open Sans', size=10, color=fg or '000000')
        nm.alignment = Alignment(horizontal='left', vertical='center', indent=1); nm.border = box
    ws.merge_cells(f'G{6+len(compact_rows)}:I{6+len(compact_rows)}')
    lc = ws.cell(row=6+len(compact_rows), column=7, value='See the "Color Key" tab for full legend')
    lc.hyperlink = "#'Color Key'!A1"
    lc.font = Font(name='Open Sans', size=9, italic=True, color=UH_GREEN, underline='single')
    lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 4
print("[P2] compact legends on chapter tabs")

# =============================================================================
# Phase 4 resource tabs (inserted BEFORE Welcome so we don't juggle indices)
# =============================================================================
def new_tab(name, pos, title, subtitle):
    if name in wb.sheetnames: del wb[name]
    ws = wb.create_sheet(name, pos)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 62
    ws.column_dimensions['E'].width = 2
    ws.sheet_properties.tabColor = UH_GREEN
    ws.merge_cells('B2:D2')
    b = ws['B2']; b.value = title
    b.font = Font(name='Open Sans', size=20, bold=True, color=UH_WHITE)
    b.fill = PatternFill('solid', start_color=UH_GREEN)
    b.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 38
    ws.merge_cells('B3:D3')
    s = ws['B3']; s.value = subtitle
    s.font = Font(name='Open Sans', size=11, italic=True, color=GRAY_TEXT)
    s.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[3].height = 20
    ws['D4'].value = '← Back to Welcome'
    ws['D4'].hyperlink = "#'Welcome'!A1"
    ws['D4'].font = Font(name='Open Sans', size=10, italic=True, color=UH_GREEN, underline='single')
    ws['D4'].alignment = Alignment(horizontal='right', vertical='center')
    return ws

def section_header(ws, row, text):
    ws.merge_cells(f'B{row}:D{row}')
    c = ws[f'B{row}']; c.value = text
    c.font = Font(name='Open Sans', size=14, bold=True, color=UH_GREEN)
    c.alignment = Alignment(horizontal='left', vertical='bottom', indent=1)
    c.border = bottom_rule
    ws.row_dimensions[row].height = 28

def table_header(ws, row, cols_labels):
    for col, label in cols_labels:
        c = ws[f'{col}{row}']; c.value = label
        c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
        c.fill = PatternFill('solid', start_color=UH_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = box
    ws.row_dimensions[row].height = 22

def text_row(ws, row, cells, *, zebra=False, bold_first=False, font_sizes=None, colors=None, wrap=True):
    font_sizes = font_sizes or [11,11,10]
    colors = colors or ['000000','000000',GRAY_TEXT]
    cols = ['B','C','D']
    for i,val in enumerate(cells):
        if val is None: continue
        c = ws[f'{cols[i]}{row}']; c.value = val
        c.font = Font(name='Open Sans', size=font_sizes[i], bold=(bold_first and i==0), color=colors[i])
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=wrap)
        c.border = box
        if zebra: c.fill = PatternFill('solid', start_color=LIGHT_GREEN)

# --- Using This Workbook -----------------------------------------------------
ws = new_tab('Using This Workbook', 1, 'Using This Workbook',
             'A short tour of the navigation and conventions.')
r = 5
section_header(ws, r, '1 · Navigation basics'); r += 1
table_header(ws, r, [('B','Action'),('C','How')]); r += 1
nav_rows = [
    ('Next tab',          'Ctrl + Page Down'),
    ('Previous tab',      'Ctrl + Page Up'),
    ('Jump to a named cell','Type the name in the Name Box (top-left of the formula bar)'),
    ('Jump to a tab',     'Right-click the tab-scroll arrows (bottom-left) for the full list'),
    ('Find anywhere',     'Ctrl + F, or Ctrl + Shift + F to search all tabs'),
    ('Recalculate',       'F9 (full) or Shift + F9 (current sheet)'),
    ('Edit a cell',       'F2'),
    ('Cancel an edit',    'Esc · Undo = Ctrl + Z'),
]
for i,(a,b) in enumerate(nav_rows):
    text_row(ws, r, [a,b], zebra=(i%2==1), bold_first=True); r += 1

r += 2; section_header(ws, r, '2 · Color coding (recap)'); r += 1
table_header(ws, r, [('B','Cell type'),('C','Color'),('D','What it means')]); r += 1
legend = [
    ('Input',         INPUT_FILL,   None,       'Hardcoded value — safe to change to explore scenarios.'),
    ('Formula',       FORMULA_FILL, FORMULA_FG, 'Calculated. Blue text. Do not edit.'),
    ('Array formula', FORMULA_FILL, ARRAY_FG,   'Dynamic/array. Purple text. Do not edit.'),
    ('Linked cell',   FORMULA_FILL, LINK_FG,    'Pulls from another tab. Green text. Do not edit.'),
]
for i,(label, fill, fg, meaning) in enumerate(legend):
    a = ws.cell(row=r, column=2, value=label); a.font = Font(name='Open Sans', size=11, bold=True); a.alignment = Alignment(horizontal='left', indent=1, vertical='center'); a.border = box
    b = ws.cell(row=r, column=3, value='sample'); b.fill = PatternFill('solid', start_color=fill)
    b.font = Font(name='Open Sans', size=11, color=(fg or '000000')); b.alignment = Alignment(horizontal='center', vertical='center'); b.border = box
    d = ws.cell(row=r, column=4, value=meaning); d.font = Font(name='Open Sans', size=10, color=GRAY_TEXT); d.alignment = Alignment(horizontal='left', indent=1, vertical='center', wrap_text=True); d.border = box
    if i%2==1:
        for col in (2,4):
            ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1

r += 2; section_header(ws, r, '3 · Hidden tabs'); r += 1
text_row(ws, r, ['Why?', 'Raw ETF price-data tabs (GBTC, GLD, IEF, SPY, EWJ, EZU, EWA, correl, data) are hidden to keep the tab strip clean.', None])
r += 1
text_row(ws, r, ['Unhide', 'Right-click any tab → Unhide → pick the tab you want to inspect.', None], zebra=True)
r += 1
text_row(ws, r, ['Why so much price data?', 'Chapter 15 uses these to compute returns, volatility, correlations, and the Sharpe ratio.', None])
r += 1

r += 2; section_header(ws, r, '4 · Before you experiment'); r += 1
rows = [
    ('Save a copy','File → Save As before making scenario changes. The master should stay clean.'),
    ('Use named ranges','Most formulas use names like rate_1y_US, forward_1y_GBPUSD, call_strike. Changing the yellow cell is enough.'),
    ('Break something?','Ctrl + Z. If you saved a broken state, ask your instructor for the master file.'),
]
for i,(a,b) in enumerate(rows):
    text_row(ws, r, [a,b], zebra=(i%2==1), bold_first=True); r += 1

# --- Excel Formulas ----------------------------------------------------------
ws = new_tab('Excel Formulas', 2, 'Excel Formulas Used in This Workbook',
             'Grouped by topic, with syntax and the chapter(s) where each function appears.')
r = 5
def category(ws, r, title, rows):
    section_header(ws, r, title); r += 1
    table_header(ws, r, [('B','Function'),('C','Syntax'),('D','What it does / used in')]); r += 1
    for i,(fn, syn, desc) in enumerate(rows):
        text_row(ws, r, [fn, syn, desc], zebra=(i%2==1), bold_first=True)
        ws.row_dimensions[r].height = 34
        r += 1
    return r + 1

r = category(ws, r, '1 · Time value & FX', [
    ('PV',    '=PV(rate, nper, pmt, [fv])', 'Present value of cash flows — used throughout Chapter 6 parity proofs.'),
    ('FV',    '=FV(rate, nper, pmt, [pv])', 'Future value of a current investment at a given rate.'),
    ('PMT',   '=PMT(rate, nper, pv, [fv])', 'Periodic payment for a loan or annuity.'),
])
r = category(ws, r, '2 · Options & derivatives', [
    ('MAX / MIN', '=MAX(spot - strike, 0)',
     'Call payoff at expiration. Chapter 7 (option fx hedge, PnL charts).'),
    ('IF',        '=IF(spot>strike, spot-strike, 0)',
     'Equivalent to MAX(...,0) — explicit branch of the option payoff.'),
    ('Payoff diagram', 'Scatter with smoothed line over an array of spot prices',
     'Used for all PnL Chart * tabs.'),
])
r = category(ws, r, '3 · Statistics / risk (Chapter 15)', [
    ('AVERAGE',     '=AVERAGE(range)', 'Mean return.'),
    ('STDEV.P',     '=STDEV.P(range)', 'Population standard deviation — used for volatility.'),
    ('STDEV.S',     '=STDEV.S(range)', 'Sample standard deviation.'),
    ('CORREL',      '=CORREL(x, y)', 'Correlation between two return series. Correl ZM + UAL tab.'),
    ('COVARIANCE.P','=COVARIANCE.P(x, y)', 'Covariance — building block of portfolio variance.'),
    ('SQRT',        '=SQRT(variance)',
     'Portfolio standard deviation is √(w\'Σw); SQRT converts variance to vol.'),
    ('SUMPRODUCT',  '=SUMPRODUCT(weights, returns)',
     'Weighted mean. Used for portfolio expected return and Sharpe numerator.'),
])
r = category(ws, r, '4 · Lookup, logic, utility', [
    ('IFERROR',   '=IFERROR(value, fallback)',
     "Hides #DIV/0! at the tail of time-series ranges."),
    ('INDIRECT',  '=INDIRECT(ref_text)', 'Turn a text string into a cell reference.'),
    ('FORMULATEXT','=FORMULATEXT(reference)',
     "Print a formula as text — used on multiple tabs to show students the equation next to the number."),
    ('& (concat)', '= "rate_1y_" & country',
     'String concatenation.'),
])

# --- Named Ranges ------------------------------------------------------------
ws = new_tab('Named Ranges', 3, 'Named Ranges Guide',
             'Why this workbook uses them and how to work with them.')
r = 5
section_header(ws, r, '1 · What is a named range?'); r += 1
ws.merge_cells(f'B{r}:D{r+2}')
b = ws[f'B{r}']
b.value = ('A named range is a human-readable alias for a cell or range. Instead of =C2*(1+C1) we write '
           '=invest*(1+rate_1y_US). That makes formulas self-documenting, easier to audit, and safer to edit.')
b.font = Font(name='Open Sans', size=11, color='333333')
b.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
r += 4
section_header(ws, r, '2 · How to create one'); r += 1
steps = [
    ('1.','Select the cell (or range) you want to name.'),
    ('2.','Type the name in the Name Box at the top-left of the formula bar, then press Enter. (Or Formulas → Define Name.)'),
    ('3.','Use the name anywhere a cell reference is allowed.'),
    ('4.','To see/edit all names: Formulas → Name Manager (Ctrl + F3).'),
]
for i,(n,t) in enumerate(steps):
    a = ws.cell(row=r, column=2, value=n); a.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN); a.alignment = Alignment(horizontal='center', vertical='center')
    c = ws.cell(row=r, column=3, value=t); c.font = Font(name='Open Sans', size=11); c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    if i%2==1:
        for col in (2,3): ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1
r += 2; section_header(ws, r, '3 · Named ranges used in this workbook'); r += 1
table_header(ws, r, [('B','Name'),('C','Where'),('D','What it is')]); r += 1
names_table = [
    ('rate_1y_US',            'Interest Rate Parity!C2',          '1-year US interest rate.'),
    ('rate_1y_UK',            'Interest Rate Parity!C3',          '1-year UK interest rate.'),
    ('invest',                'Interest Rate Parity!C5',          'Investment notional (USD).'),
    ('forward_1y_GBPUSD',     'Interest Rate Parity!C6',          '1-year GBPUSD forward rate.'),
    ('contract_notional_value_payable',   'Transaction Hedging (Payables)!F6',    'Notional amount of the foreign-currency payable.'),
    ('current_spot_price_payable',        'Transaction Hedging (Payables)!F7',    'Spot FX at the time the hedge is put on.'),
    ('call_strike · call_price','Transaction Hedging (Payables)!F14:F15','Call-option strike and premium (used for hedging a payable).'),
    ('put_price',             'Transaction Hedging (Receivable!F15','Put-option premium (used for hedging a receivable).'),
    ('fv_put_outlay',         'Transaction Hedging (Receivable!F28','Future value of the premium paid.'),
    ('gbpusd_1y_scenario',    'Transaction Hedging (Receivable!C33:C45','Vector of possible 1-year GBPUSD outcomes for scenario analysis.'),
    ('recievable',            'Transaction Hedging (Receivable!F6','Notional amount of the foreign-currency receivable.'),
]
for i,row in enumerate(names_table):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True)
    ws.row_dimensions[r].height = 30
    r += 1

# --- Data Analytics ----------------------------------------------------------
ws = new_tab('Data Analytics', 4, 'Data Analytics in Excel',
             'Descriptive statistics, histograms, correlation, Sharpe ratio.')
r = 5
section_header(ws, r, '1 · Descriptive statistics'); r += 1
table_header(ws, r, [('B','Function'),('C','Formula'),('D','Use')]); r += 1
for i,row in enumerate([
    ('Mean',    '=AVERAGE(range)',                      'Arithmetic average of returns.'),
    ('Median',  '=MEDIAN(range)',                       'Middle value — robust to outliers.'),
    ('Std dev','=STDEV.P(range) / =STDEV.S(range)',     'Volatility; .P for population, .S for sample.'),
    ('Variance','=VAR.P(range)',                        'Std dev squared.'),
    ('Min/Max','=MIN(range) / =MAX(range)',             'Range bounds.'),
    ('Percentile','=PERCENTILE.INC(range, 0.95)',       '95th percentile (common in VaR).'),
]):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True); r += 1

r += 2; section_header(ws, r, '2 · Correlation'); r += 1
for i,row in enumerate([
    ('CORREL',        '=CORREL(x_returns, y_returns)',
     'Pearson correlation — see Correl ZM + UAL.'),
    ('COVARIANCE.P',  '=COVARIANCE.P(x, y)',
     'Unscaled co-movement.'),
    ('Rolling CORREL','Apply over a sliding window',
     'Correl ZM + UAL shows rolling 30-day and multi-period correlations.'),
]):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True)
    ws.row_dimensions[r].height = 34; r += 1

r += 2; section_header(ws, r, '3 · Sharpe ratio (Chapter 15)'); r += 1
ws.merge_cells(f'B{r}:D{r+2}')
p = ws[f'B{r}']
p.value = ('The Sharpe ratio measures excess return per unit of risk: '
           'Sharpe = (E[R_portfolio] − R_risk_free) / σ_portfolio. A higher Sharpe '
           'means more reward for each unit of volatility. To build it in Excel:')
p.font = Font(name='Open Sans', size=11, color='333333')
p.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
r += 4
steps = [
    ('1.', 'Compute periodic returns with (price_t / price_{t-1}) − 1, then annualize the mean and std dev.'),
    ('2.', 'Expected return = SUMPRODUCT(weights, expected_returns) across assets.'),
    ('3.', 'Portfolio volatility = SQRT(weights \' Σ weights). For two assets: SQRT(w1²σ1² + w2²σ2² + 2·w1·w2·σ1·σ2·ρ).'),
    ('4.', 'Risk-free rate = 1-year T-bill yield (see IEF tab for a proxy).'),
    ('5.', 'Sharpe = (expected − risk-free) / portfolio volatility.'),
]
for i,(n,t) in enumerate(steps):
    a = ws.cell(row=r, column=2, value=n); a.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN); a.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'C{r}:D{r}')
    c = ws.cell(row=r, column=3, value=t); c.font = Font(name='Open Sans', size=11); c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 36
    if i%2==1:
        for col in (2,3): ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1

# --- Claude for Excel --------------------------------------------------------
ws = new_tab('Claude for Excel', 5, 'Claude for Excel',
             'Using Anthropic\'s AI assistant to work with this workbook.')
r = 5
section_header(ws, r, '1 · What is it'); r += 1
ws.merge_cells(f'B{r}:D{r+2}')
b = ws[f'B{r}']
b.value = ('Claude is Anthropic\'s AI assistant. Claude for Excel lets you ask questions about your '
           'spreadsheet, have it read and reason over your data, explain formulas, and generate new '
           'analysis. It is available inside Excel as an add-in and on claude.ai where you can upload '
           'an .xlsx file and chat about it.')
b.font = Font(name='Open Sans', size=11, color='333333')
b.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
r += 4
section_header(ws, r, '2 · Useful prompts for this workbook'); r += 1
table_header(ws, r, [('B','Topic'),('C','Try asking Claude')]); r += 1
prompts = [
    ('Parity intuition','"Explain covered interest rate parity using the Interest Rate Parity tab — what happens if the forward moves 1% in either direction?"'),
    ('Option hedging','"For Transaction Hedging (Receivables), walk me through when a put option beats a forward hedge — under what spot scenarios?"'),
    ('Sharpe math','"Compute the Sharpe ratio for a 60% ZM / 40% UAL portfolio using the data in the Portfolio tab. Show your work."'),
    ('Sensitivity','"Build a two-way sensitivity table for the receivable hedge: strike price (1.30–1.40) vs implied volatility (10%–30%)."'),
    ('Translate a number','"A correlation of 0.42 between ZM and UAL — what does that mean for portfolio diversification?"'),
    ('Concept check','"What is the difference between covered and uncovered interest parity, and why does UIP often fail empirically?"'),
    ('Generate data',   '"Give me 5 years of synthetic monthly returns for a hypothetical EM equity index with 25% annualized vol and 0.3 correlation to SPY."'),
]
for i,row in enumerate(prompts):
    a = ws.cell(row=r, column=2, value=row[0]); a.font = Font(name='Open Sans', size=11, bold=True); a.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True); a.border = box
    ws.merge_cells(f'C{r}:D{r}')
    b = ws.cell(row=r, column=3, value=row[1]); b.font = Font(name='Open Sans', size=10, color='333333', italic=True); b.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True); b.border = box
    ws.row_dimensions[r].height = 40
    if i%2==1:
        a.fill = PatternFill('solid', start_color=LIGHT_GREEN)
        b.fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1
r += 2; section_header(ws, r, '3 · How to get good answers'); r += 1
tips = [
    ('Be specific about the cell or tab','Reference exact tab names and cell addresses. "Explain Interest Rate Parity!C9" beats "explain the formula".'),
    ('Give Claude the context','Attach the workbook or paste the relevant rows. Claude cannot read what you do not show it.'),
    ('Ask for the reasoning','"Show your work" or "walk me through step by step" catches mistakes.'),
    ('Verify numbers','Cross-check arithmetic against Excel itself. AI can produce plausible but wrong numbers.'),
    ('Cite your AI use','Per the course AI policy, record meaningful prompts in deliverables/prompt-log.md.'),
]
for i,(a,b) in enumerate(tips):
    text_row(ws, r, [a,b,None], zebra=(i%2==1), bold_first=True); ws.row_dimensions[r].height = 38; r += 1
r += 2; section_header(ws, r, '4 · What NOT to do'); r += 1
for i,(a,b) in enumerate([
    ('Do not paste real personal data','SSNs, account numbers, or client-identifiable info should never go into an AI chat.'),
    ('Do not outsource your learning','Use Claude to explain and check, not to do the assignment for you.'),
    ('Do not trust without verifying','Treat AI-written Excel formulas like code from a stranger — test before shipping.'),
]):
    text_row(ws, r, [a,b,None], zebra=(i%2==1), bold_first=True); ws.row_dimensions[r].height = 36; r += 1
print("[P4] resource tabs created")

# =============================================================================
# Phase 3 — Welcome tab + tab colors + hide data tabs
# =============================================================================
# Insert Welcome at position 0 (everything else shifts +1)
if 'Welcome' in wb.sheetnames: del wb['Welcome']
w = wb.create_sheet('Welcome', 0)
w.sheet_view.showGridLines = False
w.sheet_view.zoomScale = 110
w.column_dimensions['A'].width = 2
w.column_dimensions['B'].width = 8
w.column_dimensions['C'].width = 40
w.column_dimensions['D'].width = 60
w.column_dimensions['E'].width = 2
w.sheet_properties.tabColor = UH_GREEN

w.merge_cells('B2:D2')
t = w['B2']; t.value = 'International Finance Workbook'
t.font = Font(name='Open Sans', size=22, bold=True, color=UH_WHITE)
t.fill = PatternFill('solid', start_color=UH_GREEN)
t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
w.row_dimensions[2].height = 42

w.merge_cells('B3:D3')
s = w['B3']; s.value = 'UH Mānoa · Shidler College of Business · FIN 321 / BUS 629'
s.font = Font(name='Open Sans', size=11, italic=True, color=GRAY_TEXT)
s.alignment = Alignment(horizontal='left', vertical='center', indent=1)
w.row_dimensions[3].height = 20

w.merge_cells('B5:D7')
intro = w['B5']
intro.value = ('This workbook is a reference library of worked examples for international '
               'finance. Chapter 6 covers parity conditions, Chapter 7 options, Chapter 8 '
               'transaction hedging, and Chapter 15 international portfolio management and '
               'the Sharpe ratio. Click a chapter below to jump to its contents page. The '
               'Color Key tab explains which cells are safe to edit.')
intro.font = Font(name='Open Sans', size=11, color='333333')
intro.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)

w.merge_cells('B9:D9')
hx = w['B9']; hx.value = 'Chapter Navigation'
hx.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
hx.alignment = Alignment(horizontal='left', vertical='center', indent=1)
hx.border = bottom_rule

for col, text in zip(['B','C','D'], ['Chapter','Topic','What you will find']):
    c = w[f'{col}10']; c.value = text
    c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
    c.fill = PatternFill('solid', start_color=UH_GREEN)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = box

chapter_nav = [
    ('Chapter 6 International Parity',   '6 — International Parity',
     'Covered interest parity worked out step-by-step in USD and GBP.'),
    ('Chapter 7 Options',                '7 — Options & Derivatives',
     'Long/short calls and puts, option FX hedge, 11 payoff-diagram charts.'),
    ('Chapter 8 Transaction Hedging',    '8 — Transaction Hedging',
     'Hedging a foreign-currency receivable and payable across a spot-scenario grid.'),
    ('Chapter 15  Intl Portffolio Man',  '15 — International Portfolio Management',
     'Correlations, beta, portfolio variance (ZM + UAL); 4 charts.'),
    ('Sharpe',                           'Sharpe — Asset-class comparisons',
     'ETF histories (SPY, IEF, GLD, etc.), equity-only and multi-asset Sharpe, charts.'),
]
for i,(tab_name, topic, desc) in enumerate(chapter_nav):
    r = 11 + i; w.row_dimensions[r].height = 30
    b = w.cell(row=r, column=2, value=tab_name); safe = tab_name.replace("'","''")
    b.hyperlink = f"#'{safe}'!A1"
    b.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN, underline='single')
    b.alignment = Alignment(horizontal='left', vertical='center', indent=1); b.border = box
    c = w.cell(row=r, column=3, value=topic); c.font = Font(name='Open Sans', size=11); c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True); c.border = box
    d = w.cell(row=r, column=4, value=desc); d.font = Font(name='Open Sans', size=10, color=GRAY_TEXT); d.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True); d.border = box
    if i%2==0:
        for col in (2,3,4):
            w.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)

# Resources section
res_header_row = 11 + len(chapter_nav) + 1
w.merge_cells(f'B{res_header_row}:D{res_header_row}')
h = w[f'B{res_header_row}']; h.value = 'Resources'
h.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
h.alignment = Alignment(horizontal='left', vertical='center', indent=1)
h.border = bottom_rule
res_rows = [
    ('Color Key',             'Color convention & legend',           'Every cell type with an example swatch.'),
    ('Using This Workbook',   'Tour / shortcuts / hidden tabs',      'Open this first if it is your first time.'),
    ('Excel Formulas',        'Function reference by topic',         'Every function used in the workbook with syntax.'),
    ('Named Ranges',          'Concept and workbook-specific list',  'Understand rate_1y_US, call_strike, etc.'),
    ('Data Analytics',        'Stats, correlation, Sharpe ratio',    'Tools used in Chapter 15.'),
    ('Claude for Excel',      'AI assistant walkthrough',            'Prompt ideas + dos and donts.'),
]
for i,(tab, title, desc) in enumerate(res_rows):
    r = res_header_row + 1 + i; w.row_dimensions[r].height = 26
    b = w.cell(row=r, column=2, value=tab); safe = tab.replace("'","''")
    b.hyperlink = f"#'{safe}'!A1"
    b.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN, underline='single')
    b.alignment = Alignment(horizontal='left', vertical='center', indent=1); b.border = box
    c = w.cell(row=r, column=3, value=title); c.font = Font(name='Open Sans', size=11); c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True); c.border = box
    d = w.cell(row=r, column=4, value=desc); d.font = Font(name='Open Sans', size=10, color=GRAY_TEXT); d.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True); d.border = box
    if i%2==0:
        for col in (2,3,4):
            w.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)

# Getting started
gs_row = res_header_row + len(res_rows) + 2
w.merge_cells(f'B{gs_row}:D{gs_row}')
gs = w[f'B{gs_row}']; gs.value = 'Getting started'
gs.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
gs.alignment = Alignment(horizontal='left', vertical='center', indent=1)
gs.border = bottom_rule
tips = [
    ('1.', 'Open the Color Key tab',                'The color legend explains which cells are safe to edit.'),
    ('2.', 'Pick a chapter tab from the list above','Each chapter page is a jump table to that chapter\'s examples.'),
    ('3.', 'Change a yellow input cell',            'Watch how every gray/blue formula cell updates.'),
    ('4.', 'Never edit a gray (formula) cell',      'If you accidentally do, press Ctrl+Z to undo.'),
]
for i,(num, title, body) in enumerate(tips):
    r = gs_row + 1 + i; w.row_dimensions[r].height = 22
    a = w.cell(row=r, column=2, value=num); a.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN); a.alignment = Alignment(horizontal='center', vertical='center')
    t1 = w.cell(row=r, column=3, value=title); t1.font = Font(name='Open Sans', size=11, bold=True); t1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    t2 = w.cell(row=r, column=4, value=body); t2.font = Font(name='Open Sans', size=10, color=GRAY_TEXT); t2.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)

foot_row = gs_row + 1 + len(tips) + 2
w.merge_cells(f'B{foot_row}:D{foot_row}')
foot = w[f'B{foot_row}']; foot.value = '→ Open the Color Key tab'
foot.hyperlink = "#'Color Key'!A1"
foot.font = Font(name='Open Sans', size=11, italic=True, color=UH_GREEN, underline='single')
foot.alignment = Alignment(horizontal='left', vertical='center', indent=1)
w.freeze_panes = 'A5'

print("[P3] Welcome tab created")

# Tab colors by chapter (use final sheet order AFTER Welcome insertion)
order = wb.sheetnames
# chapter groups are defined on the ORIGINAL chapter/section tabs
section_indices = [order.index(s[0]) for s in CHAPTER_MAP]
def chapter_of(idx):
    if idx < section_indices[0]:
        return None
    for i in range(len(section_indices)):
        start = section_indices[i]
        end = section_indices[i+1] if i+1 < len(section_indices) else len(order)
        if start <= idx < end:
            return CH_KEYS[i]
    return None
colored = 0
for i, nm in enumerate(order):
    if nm in ('Welcome','Color Key','Using This Workbook','Excel Formulas','Named Ranges','Data Analytics','Claude for Excel'):
        continue
    sheet = wb[nm]
    if isinstance(sheet, Chartsheet): continue
    b = chapter_of(i)
    if b:
        sheet.sheet_properties.tabColor = CHAPTER_TAB_COLORS[b]
        colored += 1
print(f"[P3] colored tabs: {colored}")

# Hide raw data tabs
hidden_count = 0
for sn in HIDE_TABS:
    if sn in wb.sheetnames:
        wb[sn].sheet_state = 'hidden'
        hidden_count += 1
print(f"[P3] hidden: {hidden_count}")

wb.save(SRC)
print("openpyxl save complete.")

# =============================================================================
# POST-SAVE FIXUP
# =============================================================================
tmp = SRC.with_suffix('.xlsx.tmp')
with zipfile.ZipFile(SRC, 'r') as zin:
    names = zin.namelist()
    buf = {n: zin.read(n) for n in names}

_rels_xml = buf['xl/_rels/workbook.xml.rels'].decode('utf-8')
_ct_xml   = buf['[Content_Types].xml'].decode('utf-8')
rel_target = {}
for r in re.finditer(r'<Relationship\b[^>]*?>', _rels_xml):
    tag = r.group(0)
    im = re.search(r'Id="([^"]+)"', tag); tm = re.search(r'Target="([^"]+)"', tag)
    if im and tm: rel_target[im.group(1)] = tm.group(1)
ct_type = {}
for o in re.finditer(r'<Override\b[^>]*?>', _ct_xml):
    tag = o.group(0)
    pm = re.search(r'PartName="([^"]+)"', tag); cm = re.search(r'ContentType="([^"]+)"', tag)
    if pm and cm: ct_type[pm.group(1)] = cm.group(1)

_wb_xml = buf['xl/workbook.xml'].decode('utf-8')
sheet_flags = []
for m in re.finditer(r'<sheet\b[^>]*?>', _wb_xml):
    tag = m.group(0)
    rm = re.search(r'r:id="([^"]+)"', tag)
    if not rm: sheet_flags.append(False); continue
    tgt = rel_target.get(rm.group(1), '')
    part = tgt if tgt.startswith('/') else ('/xl/' + tgt.lstrip('/'))
    sheet_flags.append('chart' in ct_type.get(part, '').lower())
new_ws_only_to_full = [i for i,ch in enumerate(sheet_flags) if not ch]
print(f"Sheets: {len(sheet_flags)} total, {sum(sheet_flags)} charts, "
      f"{len(new_ws_only_to_full)} worksheets")

def remap(m):
    idx = int(m.group(1))
    return f'localSheetId="{new_ws_only_to_full[idx]}"' if idx < len(new_ws_only_to_full) else m.group(0)
_wb_xml = re.sub(r'localSheetId="(\d+)"', remap, _wb_xml)
op = r'<definedName\b(?![^>]*localSheetId=)[^>]*>[^<]*(?:\[1\]|#REF!)[^<]*</definedName>'
orphans = re.findall(op, _wb_xml)
_wb_xml = re.sub(op, '', _wb_xml)
print(f"Orphan globals removed: {len(orphans)}")
buf['xl/workbook.xml'] = _wb_xml.encode('utf-8')

styles_xml = buf['xl/styles.xml'].decode('utf-8')
for old in ('Calibri','Arial','Inconsolata','Roboto'):
    styles_xml = re.sub(rf'<name\s+val="{old}"\s*/>', '<name val="Open Sans"/>', styles_xml)
buf['xl/styles.xml'] = styles_xml.encode('utf-8')

with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
    for n in names:
        zout.writestr(n, buf[n])
shutil.move(str(tmp), str(SRC))
print("Post-save fixups complete.")
