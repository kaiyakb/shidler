"""Phase 4: pedagogical resource tabs
Adds five reference tabs right after Color Key:
    1. Using This Workbook  — student onboarding / shortcuts / hidden tabs
    2. Excel Formulas       — categorized function reference
    3. Named Ranges         — concept + workbook-specific conventions
    4. Data Analytics       — stats, histograms, regression primer
    5. Claude for Excel     — AI assistant walkthrough + prompt ideas

Also updates the Welcome tab's navigation to point to these resources.
"""
import re, shutil, zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = Path(r'C:/GitHub/shidler/docs/spreadsheets/Corporate Finance Master Spreadsheets.xlsx')

UH_GREEN    = '024731'
UH_WHITE    = 'FFFFFF'
LIGHT_GREEN = 'E6EEEA'
INPUT_FILL  = 'FFF2CC'
FORMULA_FILL= 'F3F3F3'
FORMULA_FG  = '0000FF'
ARRAY_FG    = '9900FF'
LINK_FG     = '006100'
GRAY_TEXT   = '595959'

thin = Side(style='thin', color='BFBFBF')
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
bottom_rule = Border(bottom=Side(style='medium', color=UH_GREEN))

wb = load_workbook(SRC, data_only=False)

# Drop any pre-existing copies so the script is idempotent
for n in ('Using This Workbook','Excel Formulas','Named Ranges','Data Analytics','Claude for Excel'):
    if n in wb.sheetnames:
        del wb[n]

def new_tab(name, pos, title, subtitle):
    """Create a resource tab with the standard banner and return the sheet."""
    ws = wb.create_sheet(name, pos)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 62
    ws.column_dimensions['E'].width = 2
    ws.sheet_properties.tabColor = UH_GREEN

    # Banner
    ws.merge_cells('B2:D2')
    b = ws['B2']
    b.value = title
    b.font = Font(name='Open Sans', size=20, bold=True, color=UH_WHITE)
    b.fill = PatternFill('solid', start_color=UH_GREEN)
    b.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 38

    ws.merge_cells('B3:D3')
    s = ws['B3']
    s.value = subtitle
    s.font = Font(name='Open Sans', size=11, italic=True, color=GRAY_TEXT)
    s.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[3].height = 20

    # Back-to-Welcome link (row 4, right side)
    ws['D4'].value = '← Back to Welcome'
    ws['D4'].hyperlink = "#'Welcome'!A1"
    ws['D4'].font = Font(name='Open Sans', size=10, italic=True, color=UH_GREEN, underline='single')
    ws['D4'].alignment = Alignment(horizontal='right', vertical='center')

    return ws

def section_header(ws, row, text):
    """Write a section header spanning B:D with a UH-green bottom rule."""
    ws.merge_cells(f'B{row}:D{row}')
    c = ws[f'B{row}']
    c.value = text
    c.font = Font(name='Open Sans', size=14, bold=True, color=UH_GREEN)
    c.alignment = Alignment(horizontal='left', vertical='bottom', indent=1)
    c.border = bottom_rule
    ws.row_dimensions[row].height = 28

def table_header(ws, row, cols_and_labels):
    """Place a colored table header row; cols_and_labels is list[(col, label)]."""
    for col, label in cols_and_labels:
        c = ws[f'{col}{row}']
        c.value = label
        c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
        c.fill = PatternFill('solid', start_color=UH_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = box
    ws.row_dimensions[row].height = 22

def text_row(ws, row, cells, *, zebra=False, wrap=True, bold_first=False, font_sizes=None, colors=None):
    """Write a data row across columns B/C/D. `cells` is a list (len<=3) of
    string values for B, C, D (None to skip). Applies consistent styling."""
    font_sizes = font_sizes or [11, 11, 10]
    colors     = colors or ['000000', '000000', GRAY_TEXT]
    cols = ['B','C','D']
    for i, val in enumerate(cells):
        if val is None: continue
        c = ws[f'{cols[i]}{row}']
        c.value = val
        c.font = Font(name='Open Sans', size=font_sizes[i],
                      bold=(bold_first and i == 0), color=colors[i])
        c.alignment = Alignment(horizontal='left', vertical='center',
                                indent=1, wrap_text=wrap)
        c.border = box
        if zebra:
            c.fill = PatternFill('solid', start_color=LIGHT_GREEN)

# ---------------------------------------------------------------------------
# 1) Using This Workbook
# ---------------------------------------------------------------------------
ws = new_tab('Using This Workbook', 2,
             'Using This Workbook',
             'A short tour of the navigation and conventions.')

r = 5
section_header(ws, r, '1 · Navigation basics')
r += 1
table_header(ws, r, [('B','Action'),('C','How')])
r += 1
nav_rows = [
    ('Next tab',           'Ctrl + Page Down'),
    ('Previous tab',       'Ctrl + Page Up'),
    ('Jump to a named cell','Type the name in the Name Box (top-left of the formula bar)'),
    ('Jump to a tab',      'Right-click the tab-scroll arrows (bottom-left) for the full list'),
    ('Find anywhere',      'Ctrl + F, or Ctrl + Shift + F to search all tabs'),
    ('Recalculate',        'F9 (full) or Shift + F9 (current sheet)'),
    ('Edit a cell',        'F2'),
    ('Cancel an edit',     'Esc · Undo = Ctrl + Z'),
]
for i,(a,b) in enumerate(nav_rows):
    text_row(ws, r, [a, b], zebra=(i%2==1), bold_first=True)
    r += 1

r += 2
section_header(ws, r, '2 · Color coding (recap)')
r += 1
table_header(ws, r, [('B','Cell type'),('C','Color'),('D','What it means')])
r += 1
legend = [
    ('Input',         INPUT_FILL,   None,       'Hardcoded value — safe to change to explore scenarios.'),
    ('Formula',       FORMULA_FILL, FORMULA_FG, 'Calculated. Blue text. Do not edit.'),
    ('Array formula', FORMULA_FILL, ARRAY_FG,   'Dynamic named-range lookup. Purple text. Do not edit.'),
    ('Linked cell',   FORMULA_FILL, LINK_FG,    'Pulls from another tab. Green text. Do not edit.'),
]
for i,(label, fill, fg, meaning) in enumerate(legend):
    a = ws.cell(row=r, column=2, value=label); a.font = Font(name='Open Sans', size=11, bold=True); a.alignment = Alignment(horizontal='left', indent=1, vertical='center'); a.border = box
    b = ws.cell(row=r, column=3, value='sample'); b.fill = PatternFill('solid', start_color=fill)
    b.font = Font(name='Open Sans', size=11, color=(fg or '000000')); b.alignment = Alignment(horizontal='center', vertical='center'); b.border = box
    d = ws.cell(row=r, column=4, value=meaning); d.font = Font(name='Open Sans', size=10, color=GRAY_TEXT); d.alignment = Alignment(horizontal='left', indent=1, vertical='center', wrap_text=True); d.border = box
    if i % 2 == 1:
        for col in (2,3,4):
            if col != 3:  # don't zebra over the swatch
                ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1

r += 2
section_header(ws, r, '3 · Hidden tabs')
r += 1
text_row(ws, r, ['Why?',
                 'Raw ETF price-data tabs (GBTC, GLD, IEF, SPY, EWJ, EZU, EWA, correl, data) are hidden to keep the tab strip clean.',
                 None], wrap=True)
r += 1
text_row(ws, r, ['Unhide',
                 'Right-click any tab → Unhide → pick the tab you want to inspect.',
                 None], zebra=True, wrap=True)
r += 1
text_row(ws, r, ['Why so much price data?',
                 "Chapter 11 uses these to calculate historical volatilities, correlations, and compare asset classes. You don't need to edit them to use the chapter's charts.",
                 None], wrap=True)
r += 1

r += 2
section_header(ws, r, '4 · Before you experiment')
r += 1
rows = [
    ('Save a copy', 'File → Save As before making scenario changes. The master workbook should stay clean.'),
    ('Use named ranges','Most formulas use names like rate, time, WACC. Changing yellow cells is enough — the named ranges update automatically.'),
    ('Break something?','Ctrl + Z. If you saved a broken state, ask your instructor for the master file.'),
]
for i,(a,b) in enumerate(rows):
    text_row(ws, r, [a, b], zebra=(i%2==1), bold_first=True)
    r += 1
print('Built: Using This Workbook')

# ---------------------------------------------------------------------------
# 2) Excel Formulas
# ---------------------------------------------------------------------------
ws = new_tab('Excel Formulas', 3,
             'Excel Formulas Used in This Workbook',
             'Grouped by topic, with syntax and the chapter(s) where each function appears.')

r = 5
def category_section(ws, r, title, rows):
    section_header(ws, r, title); r += 1
    table_header(ws, r, [('B','Function'),('C','Syntax'),('D','What it does / used in')]); r += 1
    for i,(fn, syn, desc) in enumerate(rows):
        text_row(ws, r, [fn, syn, desc], zebra=(i%2==1), bold_first=True,
                 font_sizes=[11,11,10])
        ws.row_dimensions[r].height = 34
        r += 1
    return r + 1

r = category_section(ws, r, '1 · Time Value of Money', [
    ('PV',    '=PV(rate, nper, pmt, [fv], [type])',
     'Present value of a series of cash flows. Chapter 5 (multiple tabs), Chapter 6 (bond pricing).'),
    ('FV',    '=FV(rate, nper, pmt, [pv], [type])',
     'Future value. Chapter 5 (FV Manhattan, Retirement Annuities Due).'),
    ('PMT',   '=PMT(rate, nper, pv, [fv], [type])',
     'Constant periodic payment required for a loan or annuity. Chapter 5 (Mortgages, Retirement).'),
    ('NPER',  '=NPER(rate, pmt, pv, [fv], [type])',
     'Number of periods to reach a future value given periodic payments. Chapter 5.'),
    ('RATE',  '=RATE(nper, pmt, pv, [fv], [type], [guess])',
     'Periodic interest rate. Chapter 5 (PV calc rate).'),
    ('EFFECT','=EFFECT(nominal_rate, npery)',
     'Effective annual rate from a stated (nominal) rate. Chapter 5 (Effective Interest Rates).'),
    ('NOMINAL','=NOMINAL(effect_rate, npery)',
     'Nominal annual rate from an effective rate. Chapter 5.'),
])

r = category_section(ws, r, '2 · Investment analysis', [
    ('NPV',   '=NPV(rate, value1, value2, ...)',
     'Net present value of a stream of periodic cash flows (assumes equal intervals, first cash flow at end of period 1). Chapter 8.'),
    ('IRR',   '=IRR(values, [guess])',
     'Internal rate of return of a cash-flow stream. Chapter 8 (NPV (IRR)).'),
    ('MIRR',  '=MIRR(values, finance_rate, reinvest_rate)',
     'Modified IRR — accounts for different financing and reinvestment rates.'),
    ('XNPV',  '=XNPV(rate, values, dates)',
     'NPV for irregularly-timed cash flows.'),
    ('XIRR',  '=XIRR(values, dates, [guess])',
     'IRR for irregularly-timed cash flows.'),
])

r = category_section(ws, r, '3 · Statistics / Risk', [
    ('AVERAGE',   '=AVERAGE(range)', 'Arithmetic mean.'),
    ('STDEV.P',   '=STDEV.P(range)', 'Population standard deviation. Chapter 11 (Risk S&P Historical, Asset Classes).'),
    ('STDEV.S',   '=STDEV.S(range)', 'Sample standard deviation.'),
    ('VAR.P / VAR.S','=VAR.P(range)', 'Variance (population / sample).'),
    ('CORREL',    '=CORREL(array1, array2)', 'Pearson correlation coefficient. Chapter 11 (Risk Correl ZM + UAL, correl).'),
    ('COVARIANCE.P','=COVARIANCE.P(array1, array2)', 'Population covariance.'),
    ('SLOPE',     '=SLOPE(known_y, known_x)', 'Regression slope — beta in Chapter 12 (Beta TSLA SPX).'),
    ('INTERCEPT', '=INTERCEPT(known_y, known_x)', 'Regression intercept — alpha in CAPM.'),
    ('RSQ',       '=RSQ(known_y, known_x)', 'Coefficient of determination.'),
    ('LINEST',    '=LINEST(known_y, known_x, [const], [stats])', 'Full linear regression output (returns an array).'),
    ('FREQUENCY', '=FREQUENCY(data_array, bins_array)', 'Histogram bin counts. Chapter 11 (Risk S&P Histogram).'),
])

r = category_section(ws, r, '4 · Lookup, logic, and utility', [
    ('INDIRECT', '=INDIRECT(ref_text)',
     'Turn a text string into a cell reference. Chapter 3/4 (Ratios tab uses this to resolve dynamic named ranges).'),
    ('INDEX / MATCH','=INDEX(array, MATCH(lookup, lookup_array, 0))',
     'Flexible alternative to VLOOKUP/XLOOKUP.'),
    ('XLOOKUP',  '=XLOOKUP(lookup, lookup_array, return_array, [if_not_found])',
     'Modern lookup (Excel 365/2021+).'),
    ('IF / IFS', '=IF(test, if_true, if_false)',
     'Conditional value.'),
    ('IFERROR',  '=IFERROR(value, value_if_error)',
     "Return a fallback when a formula errors. Used to hide #DIV/0! at the end of time-series ranges."),
    ('SUMPRODUCT','=SUMPRODUCT(arr1, arr2, ...)',
     'Multiply-and-sum. Used for weighted averages (e.g., WACC).'),
    ('FORMULATEXT','=FORMULATEXT(reference)',
     'Return the formula text of a referenced cell as a string — useful for showing students the equation alongside its result. Appears next to every formula on the Ratios tab and Bonds tabs.'),
    ('& (concat)', '= "text_" & cell & "_" & year',
     'String concatenation. Used in the Ratios tab to build dynamic named-range references like "BAL_equity_shareholders_" & yearStart.'),
])
print('Built: Excel Formulas')

# ---------------------------------------------------------------------------
# 3) Named Ranges
# ---------------------------------------------------------------------------
ws = new_tab('Named Ranges', 4,
             'Named Ranges Guide',
             'Why this workbook uses them, how they are organized, and how to work with them.')

r = 5
section_header(ws, r, '1 · What is a named range?')
r += 1
ws.merge_cells(f'B{r}:D{r+2}')
b = ws[f'B{r}']
b.value = ('A named range is a human-readable alias for a cell or range. Instead of writing =C5*shares_outstanding '
           'we can write =share_price*shares_outstanding. It makes formulas self-documenting, easier to audit, and '
           'safer to edit — you can rearrange rows and the formulas still point to the right data.')
b.font = Font(name='Open Sans', size=11, color='333333')
b.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
r += 4

section_header(ws, r, '2 · How to create one')
r += 1
steps = [
    ('1.','Select the cell (or range) you want to name.'),
    ('2.','Type the name in the Name Box at the top-left of the formula bar, then press Enter. (Alternatively, Formulas → Define Name.)'),
    ('3.','Use the name anywhere a cell reference is allowed.'),
    ('4.','To see or edit all names: Formulas → Name Manager (or Ctrl + F3).'),
]
for i,(n,t) in enumerate(steps):
    a = ws.cell(row=r, column=2, value=n); a.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN); a.alignment = Alignment(horizontal='center', vertical='center')
    b = ws.cell(row=r, column=3, value=t); b.font = Font(name='Open Sans', size=11); b.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    if i%2==1:
        for col in (2,3): ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1

r += 2
section_header(ws, r, '3 · Scope: workbook-level vs sheet-level')
r += 1
table_header(ws, r, [('B','Scope'),('C','When to use'),('D','Example in this workbook')])
r += 1
scopes = [
    ('Workbook','Values that are global — e.g., company-wide inputs.',
     'WACC, beta, tax, growth_rate (Chapter 13 Value Business tabs).'),
    ('Sheet',   'Values that only make sense on one tab — avoids name collisions.',
     'rate, time, present_value, future_value on each TVM example (scoped to its own sheet so every Chapter 5 tab has its own).'),
]
for i,row in enumerate(scopes):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True)
    r += 1

r += 2
section_header(ws, r, '4 · Naming conventions used in this workbook')
r += 1
table_header(ws, r, [('B','Prefix'),('C','Meaning'),('D','Sample names')])
r += 1
conventions = [
    ('BAL_',    'Balance Sheet line item',    'BAL_equity_shareholders_2020, BAL_assets_current_2019'),
    ('INC_',    'Income Statement line item','INC_sales, INC_net, INC_interest_expense'),
    ('CASH_',   'Cash Flow Statement item', 'CASH_operating, CASH_investing'),
    ('startYear_ · currentYear_','Time-period accessor pattern', 'startYear_equity, currentYear_assets_total'),
    ('avg_',    'Average across two periods','avg_equity, avg_total_assets, avg_total_capitalization'),
    ('rate, time, pv, fv, pmt','Per-sheet TVM parameters (sheet-scoped)','Each Chapter 5 tab has its own.'),
    ('lt_debt_* · wgt_*','Chapter 13 valuation inputs/weights','lt_debt_ytm, wgt_debt, wgt_equity, lt_debt_mv'),
]
for i,row in enumerate(conventions):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True)
    ws.row_dimensions[r].height = 30
    r += 1

r += 2
section_header(ws, r, '5 · Pro tip')
r += 1
ws.merge_cells(f'B{r}:D{r+1}')
t = ws[f'B{r}']
t.value = ('On the Ratios tab, column C shows =FORMULATEXT(B?) — which prints the formula in column B as '
           'text. This is how you can see the named-range math (e.g., market_capitalization - currentYear_equity) '
           'right next to the numerical answer. Use this trick on any tab to audit your own formulas.')
t.font = Font(name='Open Sans', size=11, color='333333')
t.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
print('Built: Named Ranges')

# ---------------------------------------------------------------------------
# 4) Data Analytics
# ---------------------------------------------------------------------------
ws = new_tab('Data Analytics', 5,
             'Data Analytics in Excel',
             'Descriptive statistics, histograms, regression — the tools used in Chapters 11–12.')

r = 5
section_header(ws, r, '1 · Descriptive statistics')
r += 1
table_header(ws, r, [('B','Function'),('C','Formula'),('D','Use')])
r += 1
stats = [
    ('Mean',    '=AVERAGE(range)',             'Arithmetic average of returns.'),
    ('Median',  '=MEDIAN(range)',              'Middle value — robust to outliers.'),
    ('Std dev','=STDEV.P(range) / STDEV.S(range)','Volatility; population (.P) when you have the full history, sample (.S) when inferring from a subset.'),
    ('Variance','=VAR.P(range) / VAR.S(range)','Std dev squared.'),
    ('Min / Max','=MIN(range) / =MAX(range)','Range bounds.'),
    ('Percentile','=PERCENTILE.INC(range, 0.95)','95th percentile — used in VaR.'),
    ('Skewness','=SKEW(range)',                 'Distribution asymmetry.'),
    ('Kurtosis','=KURT(range)',                  'Tail fatness.'),
]
for i,row in enumerate(stats):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True); r += 1

r += 2
section_header(ws, r, '2 · Building a histogram')
r += 1
ws.merge_cells(f'B{r}:D{r+1}')
p = ws[f'B{r}']
p.value = ('A histogram counts how often returns fall inside "bins" (e.g., ≤ –10%, –10% to –5%, …). '
           'Three Excel paths — pick whichever is available to you:')
p.font = Font(name='Open Sans', size=11, color='333333')
p.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
r += 3

methods = [
    ('Method A · FREQUENCY',
     ('Create a column of bin upper edges (e.g., -10%, -5%, 0%, 5%, 10%). '
      'Select as many cells as you have bins, type =FREQUENCY(data, bins), then press Ctrl+Shift+Enter to enter '
      'as an array formula. The cells now hold the counts. Plot those with a Column chart. '
      'This is exactly how the Risk S&P Histogram tab was built.')),
    ('Method B · Insert > Chart > Histogram (Excel 2016+)',
     ('Select your data column. Insert → Chart → Histogram. Excel auto-bins — right-click an x-axis label to tweak '
      'bin width. Fastest but gives you less control over bin boundaries.')),
    ('Method C · Analysis ToolPak',
     ('File → Options → Add-Ins → Analysis ToolPak → Go → check the box. Data tab now has Data Analysis → Histogram. '
      'Lets you specify bin boundaries and output location; produces a frequency table and a chart.')),
]
for i,(title, body) in enumerate(methods):
    a = ws.cell(row=r, column=2, value=title); a.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN); a.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True); a.border = box
    ws.merge_cells(f'C{r}:D{r}')
    b = ws.cell(row=r, column=3, value=body); b.font = Font(name='Open Sans', size=10, color='333333'); b.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True); b.border = box
    ws.row_dimensions[r].height = 56
    if i%2==1:
        a.fill = PatternFill('solid', start_color=LIGHT_GREEN)
        b.fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1

r += 2
section_header(ws, r, '3 · Correlation & covariance')
r += 1
corr_rows = [
    ('CORREL',       '=CORREL(asset1_returns, asset2_returns)',
     'Measures co-movement, scaled to [–1, +1]. See the correl tab (hidden) and Risk Correl ZM + UAL.'),
    ('COVARIANCE.P', '=COVARIANCE.P(asset1, asset2)',
     'Unscaled co-movement — useful when constructing variance-covariance matrices for portfolios.'),
    ('Rolling correlation','Apply CORREL over a sliding window (e.g., 30 days).',
     'Risk Correl ZM + UAL uses rolling 30-day CORREL to show how correlations change over time.'),
]
for i,row in enumerate(corr_rows):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True); ws.row_dimensions[r].height = 34; r += 1

r += 2
section_header(ws, r, '4 · Regression & beta (CAPM)')
r += 1
reg_rows = [
    ('SLOPE / INTERCEPT',
     '=SLOPE(stock_returns, market_returns)',
     'Slope = beta in the CAPM. See Chapter 12 Beta TSLA SPX.'),
    ('LINEST (full output)',
     '=LINEST(y_range, x_range, TRUE, TRUE)',
     'Returns a 2-col x 5-row array: coefficients, standard errors, R², F-stat, SS. Enter as array (Ctrl+Shift+Enter).'),
    ('RSQ',
     '=RSQ(stock_returns, market_returns)',
     'R² — how much of the stock\'s variance is explained by the market.'),
    ('Scatter + trendline',
     'Chart type: Scatter. Right-click any point → Add Trendline → check Display Equation and Display R² on chart.',
     'Visual check that your slope matches the chart. Beta Scatter Chart TSLA does exactly this.'),
]
for i,row in enumerate(reg_rows):
    text_row(ws, r, list(row), zebra=(i%2==1), bold_first=True); ws.row_dimensions[r].height = 38; r += 1

print('Built: Data Analytics')

# ---------------------------------------------------------------------------
# 5) Claude for Excel
# ---------------------------------------------------------------------------
ws = new_tab('Claude for Excel', 6,
             'Claude for Excel',
             'Using Anthropic\'s AI assistant to work with this workbook.')

r = 5
section_header(ws, r, '1 · What is it')
r += 1
ws.merge_cells(f'B{r}:D{r+2}')
b = ws[f'B{r}']
b.value = ('Claude is Anthropic\'s AI assistant. Claude for Excel lets you ask questions about your spreadsheet, '
           'have it read and reason over your data, explain formulas, and generate new analysis. It is available '
           'inside Excel as an add-in and on claude.ai where you can upload an .xlsx file and chat about it.')
b.font = Font(name='Open Sans', size=11, color='333333')
b.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
r += 4

section_header(ws, r, '2 · Useful prompts for this workbook')
r += 1
table_header(ws, r, [('B','Topic'),('C','Try asking Claude')])
r += 1
prompts = [
    ('Understand a formula',
     '"Explain what the formula in cell B20 of the Value Business (step 4) tab does, in plain English."'),
    ('Audit a model',
     '"Audit the Value Business (step 4) DCF. What are its key assumptions and which cells are most sensitive?"'),
    ('Sensitivity analysis',
     '"Build a two-way sensitivity table for NPV against discount rate (5%–15%) and growth rate (2%–8%)."'),
    ('Translate a number',
     '"What does a WACC of 10.8% mean for valuing this business? What would a 1% drop in WACC do to the equity value?"'),
    ('Scenario modeling',
     '"Create three scenarios — base, bull, bear — for the Stocks (Non-Constant Growth) tab with different growth rates."'),
    ('Teach a concept',
     '"Explain why Constant Growth DDM and PVGO are equivalent expressions of the same price."'),
    ('Generate test data',
     '"Give me five years of realistic monthly returns for a hypothetical biotech stock with 45% annualized volatility."'),
]
for i,row in enumerate(prompts):
    a = ws.cell(row=r, column=2, value=row[0]); a.font = Font(name='Open Sans', size=11, bold=True); a.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True); a.border = box
    ws.merge_cells(f'C{r}:D{r}')
    b = ws.cell(row=r, column=3, value=row[1]); b.font = Font(name='Open Sans', size=10, color='333333', italic=True); b.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True); b.border = box
    ws.row_dimensions[r].height = 40
    if i%2==1:
        a.fill = PatternFill('solid', start_color=LIGHT_GREEN)
        b.fill = PatternFill('solid', start_color=LIGHT_GREEN)
    r += 1

r += 2
section_header(ws, r, '3 · How to get good answers')
r += 1
tips = [
    ('Be specific about the cell or tab', 'Reference exact tab names and cell addresses. "Explain FV Manhattan!B20" beats "explain the formula".'),
    ('Give Claude the context', 'Attach the workbook, or paste the relevant rows. Claude cannot read what you do not show it.'),
    ('Ask for the reasoning', 'Prompts like "show your work" or "walk me through step by step" catch mistakes.'),
    ('Verify numbers', 'Always cross-check Claude\'s arithmetic against Excel itself. AI can and does produce plausible but wrong numbers.'),
    ('Cite your AI use', 'Per the course AI policy, record meaningful prompts in deliverables/prompt-log.md.'),
]
for i,(t,b) in enumerate(tips):
    text_row(ws, r, [t, b, None], zebra=(i%2==1), bold_first=True)
    ws.row_dimensions[r].height = 38
    r += 1

r += 2
section_header(ws, r, '4 · What NOT to do')
r += 1
warnings = [
    ('Do not paste real personal data', 'SSNs, account numbers, or client-identifiable info should never go into an AI chat.'),
    ('Do not outsource your learning', 'Use Claude to explain and check, not to do the assignment for you.'),
    ('Do not trust without verifying', 'Treat AI-written Excel formulas like code from a stranger — test before shipping.'),
]
for i,(t,b) in enumerate(warnings):
    text_row(ws, r, [t, b, None], zebra=(i%2==1), bold_first=True)
    ws.row_dimensions[r].height = 36
    r += 1

print('Built: Claude for Excel')

# ---------------------------------------------------------------------------
# Update Welcome tab navigation to point to the new resources
# ---------------------------------------------------------------------------
welcome = wb['Welcome']
# Find current navigation end (after last chapter row) and append a Resources block
# Navigation was at rows 11..19 for 9 chapters.  Add a Resources section after.
res_header_row = 21
welcome.merge_cells(f'B{res_header_row}:D{res_header_row}')
h = welcome[f'B{res_header_row}']
h.value = 'Resources'
h.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
h.alignment = Alignment(horizontal='left', vertical='center', indent=1)
h.border = bottom_rule

res_rows = [
    ('Color Key',             'Color convention & legend',            'Every cell type with an example swatch.'),
    ('Using This Workbook',   'Tour / shortcuts / hidden tabs',        'Open this first if it is your first time.'),
    ('Excel Formulas',        'Function reference by topic',            'Every function used in the workbook with syntax.'),
    ('Named Ranges',          'Concept, scope, and conventions',        'Understand the BAL_ / INC_ / CASH_ naming.'),
    ('Data Analytics',        'Stats, histograms, regression',          'Tools used in Chapters 11 and 12.'),
    ('Claude for Excel',      'AI assistant walkthrough',               'Prompt ideas + do\'s and don\'ts.'),
]
for i,(tab, title, desc) in enumerate(res_rows):
    r = res_header_row + 1 + i
    welcome.row_dimensions[r].height = 26
    b = welcome.cell(row=r, column=2, value=tab)
    safe = tab.replace("'", "''")
    b.hyperlink = f"#'{safe}'!A1"
    b.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN, underline='single')
    b.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    b.border = box
    c = welcome.cell(row=r, column=3, value=title)
    c.font = Font(name='Open Sans', size=11)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    c.border = box
    d = welcome.cell(row=r, column=4, value=desc)
    d.font = Font(name='Open Sans', size=10, color=GRAY_TEXT)
    d.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    d.border = box
    if i%2==0:
        for col in (2,3,4):
            welcome.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)

# Shift the "Getting started" section down (it was at row 21; now it collides).
# Rebuild "Getting started" at row 29 by copying the previous tips.
# But the tips were already written at rows 22-25 and the footer at 28.
# Simplest: leave the Resources where it is, but check the "Getting started"
# region and move it.  We can detect by scanning welcome rows for 'Getting started'.
# Actually — our new Resources block spans rows 21..27.  The existing Getting
# started block is at 21..25.  Overwrite conflict resolved by clearing rows
# 21..28 first.  Let me do that more cleanly:

# Undo: delete the overlap (rows 21..28 were pre-existing "Getting started" +
# footer).  Rewrite them below Resources.
# First unmerge any ranges that intersect rows 21..29.
clear_rows = list(range(21, 30))
for mr in list(welcome.merged_cells.ranges):
    if mr.min_row in clear_rows or mr.max_row in clear_rows:
        welcome.unmerge_cells(str(mr))
for rr in clear_rows:
    for cc in range(1, 8):
        cell = welcome.cell(row=rr, column=cc)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()
        cell.hyperlink = None

# Re-render Resources (rows 21..27)
welcome.merge_cells(f'B{res_header_row}:D{res_header_row}')
h = welcome[f'B{res_header_row}']
h.value = 'Resources'
h.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
h.alignment = Alignment(horizontal='left', vertical='center', indent=1)
h.border = bottom_rule
for i,(tab, title, desc) in enumerate(res_rows):
    r = res_header_row + 1 + i
    welcome.row_dimensions[r].height = 26
    b = welcome.cell(row=r, column=2, value=tab); safe=tab.replace("'","''")
    b.hyperlink = f"#'{safe}'!A1"
    b.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN, underline='single')
    b.alignment = Alignment(horizontal='left', vertical='center', indent=1); b.border = box
    c = welcome.cell(row=r, column=3, value=title); c.font = Font(name='Open Sans', size=11); c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True); c.border = box
    d = welcome.cell(row=r, column=4, value=desc); d.font = Font(name='Open Sans', size=10, color=GRAY_TEXT); d.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True); d.border = box
    if i%2==0:
        for col in (2,3,4):
            welcome.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)

# Re-render Getting started + footer at rows 29+
gs_row = 29
welcome.merge_cells(f'B{gs_row}:D{gs_row}')
gs = welcome[f'B{gs_row}']
gs.value = 'Getting started'
gs.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
gs.alignment = Alignment(horizontal='left', vertical='center', indent=1)
gs.border = bottom_rule
tips = [
    ('1.', 'Open the Color Key tab',                       'The color legend explains which cells are safe to edit.'),
    ('2.', 'Pick a chapter tab from the list above',        'Each chapter page is a jump table to that chapter\'s examples.'),
    ('3.', 'Change a yellow input cell',                    'Watch how every gray/blue formula cell updates.'),
    ('4.', 'Never edit a gray (formula) cell',              'If you accidentally do, press Ctrl+Z to undo.'),
]
for i,(num, title, body) in enumerate(tips):
    r = gs_row + 1 + i
    welcome.row_dimensions[r].height = 22
    a = welcome.cell(row=r, column=2, value=num); a.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN); a.alignment = Alignment(horizontal='center', vertical='center')
    t1 = welcome.cell(row=r, column=3, value=title); t1.font = Font(name='Open Sans', size=11, bold=True); t1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    t2 = welcome.cell(row=r, column=4, value=body); t2.font = Font(name='Open Sans', size=10, color=GRAY_TEXT); t2.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)

foot_row = gs_row + 1 + len(tips) + 2
welcome.merge_cells(f'B{foot_row}:D{foot_row}')
foot = welcome[f'B{foot_row}']
foot.value = '→ Open the Color Key tab'
foot.hyperlink = "#'Color Key'!A1"
foot.font = Font(name='Open Sans', size=11, italic=True, color=UH_GREEN, underline='single')
foot.alignment = Alignment(horizontal='left', vertical='center', indent=1)

print('Welcome tab navigation updated.')

wb.save(SRC)
print('openpyxl save complete.')

# ---------------------------------------------------------------------------
# POST-SAVE FIXUP (localSheetId remap + orphan cleanup + font normalize)
# ---------------------------------------------------------------------------
tmp = SRC.with_suffix('.xlsx.tmp')
with zipfile.ZipFile(SRC, 'r') as zin:
    names = zin.namelist()
    buf = {n: zin.read(n) for n in names}

_rels_xml = buf['xl/_rels/workbook.xml.rels'].decode('utf-8')
_ct_xml   = buf['[Content_Types].xml'].decode('utf-8')
rel_target = {}
for r in re.finditer(r'<Relationship\b[^>]*?>', _rels_xml):
    tag = r.group(0)
    im = re.search(r'Id="([^"]+)"', tag); tm = re.search(r'Target="([^"]+)"', tag)
    if im and tm: rel_target[im.group(1)] = tm.group(1)
ct_type = {}
for o in re.finditer(r'<Override\b[^>]*?>', _ct_xml):
    tag = o.group(0)
    pm = re.search(r'PartName="([^"]+)"', tag); cm = re.search(r'ContentType="([^"]+)"', tag)
    if pm and cm: ct_type[pm.group(1)] = cm.group(1)

_wb_xml = buf['xl/workbook.xml'].decode('utf-8')
sheet_flags = []
for m in re.finditer(r'<sheet\b[^>]*?>', _wb_xml):
    tag = m.group(0)
    rm = re.search(r'r:id="([^"]+)"', tag)
    if not rm:
        sheet_flags.append(False); continue
    tgt = rel_target.get(rm.group(1), '')
    part = tgt if tgt.startswith('/') else ('/xl/' + tgt.lstrip('/'))
    sheet_flags.append('chart' in ct_type.get(part, '').lower())
new_ws_only_to_full = [i for i,ch in enumerate(sheet_flags) if not ch]
print(f"Sheets: {len(sheet_flags)} total, {sum(sheet_flags)} charts, "
      f"{len(new_ws_only_to_full)} worksheets")

def remap(m):
    idx = int(m.group(1))
    return f'localSheetId="{new_ws_only_to_full[idx]}"' if idx < len(new_ws_only_to_full) else m.group(0)
_wb_xml = re.sub(r'localSheetId="(\d+)"', remap, _wb_xml)

op = r'<definedName\b(?![^>]*localSheetId=)[^>]*>[^<]*(?:\[1\]|#REF!)[^<]*</definedName>'
orphans = re.findall(op, _wb_xml)
_wb_xml = re.sub(op, '', _wb_xml)
print(f"Orphan globals removed: {len(orphans)}")
buf['xl/workbook.xml'] = _wb_xml.encode('utf-8')

styles_xml = buf['xl/styles.xml'].decode('utf-8')
for old in ('Calibri','Arial','Inconsolata','Roboto'):
    styles_xml = re.sub(rf'<name\s+val="{old}"\s*/>', '<name val="Open Sans"/>', styles_xml)
buf['xl/styles.xml'] = styles_xml.encode('utf-8')

with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
    for n in names:
        zout.writestr(n, buf[n])
shutil.move(str(tmp), str(SRC))
print('Post-save fixups complete.')
