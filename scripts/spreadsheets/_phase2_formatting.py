"""Phase 2: consistent formatting across all worksheets.
Scope (per user instruction, 2026-04-23):
  - UH brand ONLY on chapter section tabs (already done in Phase 1; add a
    compact legend to each)
  - Formula / input / output / link cell formatting made CONSISTENT across
    all sheets using the Ratios tab as the template
  - Add a master Color Key / Legend tab as sheet #1
  - Default font -> Open Sans (XML-level style edit, preserving size/bold)

Convention (derived from Ratios):
    Input (hardcoded #)        yellow fill #FFF2CC, font as-is
    Formula (calc)             gray fill #F3F3F3, blue font #0000FF
    Array formula              gray fill #F3F3F3, purple font #9900FF
    Cross-sheet linked cell    gray fill #F3F3F3, green font #006100
    Header / label / bold      left alone
"""
import re, shutil, zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.formula import ArrayFormula

SRC = Path(r'C:/GitHub/shidler/docs/spreadsheets/Corporate Finance Master Spreadsheets.xlsx')

# ---- UH & convention colors -------------------------------------------------
UH_GREEN    = '024731'
UH_WHITE    = 'FFFFFF'
LIGHT_GREEN = 'E6EEEA'
INPUT_FILL  = 'FFF2CC'   # yellow — inputs
FORMULA_FILL= 'F3F3F3'   # light gray — computed cells
FORMULA_FG  = '0000FF'   # blue — formulas
ARRAY_FG    = '9900FF'   # purple — array / named-range formulas
LINK_FG     = '006100'   # green — cross-sheet references

# Colors we'll leave alone (already intentionally styled headers)
PROTECTED_FILLS = {
    'FFD9D9D9',  # dark gray — ratios section heads
    'FF9FC5E8',  # light blue — category bands
    'FFEFEFEF',  # light-gray header band
    'FF024731',  # UH green
    'FFE6EEEA',  # chapter zebra
    'FF000000',  # black
}

# Tabs to skip entirely (data dumps & the Ratios template itself)
SKIP_COLOR_TABS = {
    'Ratios',                       # template — already canonical
    'Risk S&P Historical', 'Risk Correl ZM + UAL', 'Risk Portfolio',
    'GBTC','GLD','IEF','SPY','EWJ','EZU','EWA',
    'Asset Classes','Asset Classes wBTC',
    'Beta TSLA  SPX',
}

wb = load_workbook(SRC, data_only=False)
full_order = list(wb.sheetnames)
is_chart = [isinstance(wb[n], Chartsheet) for n in full_order]

chapter_section_tabs = {
    'Chapter 3 + 4','Chapter 5','Chapter 6','Chapter 7 Valuing Stocks',
    'Chapter 8 NPV','Chapter 11 Intro to Risk','Chapter 12 Risk, Return, Budget',
    'Chapter 13 WACC','Chapter 23 Options',
}

def safe_fill_rgb(cell):
    try:
        v = cell.fill.fgColor.rgb
        if isinstance(v, str):
            return v.upper()
    except Exception:
        pass
    return None

def has_cross_sheet_ref(formula, current_sheet):
    """Return True if formula references a different worksheet by name."""
    # patterns: 'SheetName'!A1  or  SheetName!A1
    refs = re.findall(r"'?([A-Za-z][^!'\"]*)'?!", formula)
    for s in refs:
        if s and s != current_sheet:
            return True
    return False

# ---------------------------------------------------------------------------
# 1) APPLY CONSISTENT COLORING
# ---------------------------------------------------------------------------
stats = {'inputs': 0, 'formulas': 0, 'arrays': 0, 'links': 0, 'skipped': 0}

for sn in wb.sheetnames:
    if sn in chapter_section_tabs or sn in SKIP_COLOR_TABS:
        continue
    ws = wb[sn]
    if isinstance(ws, Chartsheet):
        continue

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            # Skip bold/header-looking cells
            if cell.font.bold:
                continue
            # Skip cells whose fill is already intentionally styled (header bands)
            existing = safe_fill_rgb(cell)
            if existing and existing in PROTECTED_FILLS:
                stats['skipped'] += 1
                continue

            # Build a replacement font that preserves size/weight/italic but
            # overrides name+color per convention. Color None -> leave unchanged.
            def new_font(color_rgb):
                return Font(
                    name='Open Sans',
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    underline=cell.font.underline,
                    color=color_rgb,
                )

            is_formula = isinstance(v, str) and v.startswith('=')
            is_arrayf  = isinstance(v, ArrayFormula)

            if is_arrayf:
                cell.fill = PatternFill('solid', start_color=FORMULA_FILL)
                cell.font = new_font(ARRAY_FG)
                stats['arrays'] += 1
            elif is_formula:
                # FORMULATEXT helpers keep existing convention (blue font,
                # transparent or white fill)
                if '_xlfn.FORMULATEXT(' in v:
                    cell.font = new_font(FORMULA_FG)
                    continue
                if has_cross_sheet_ref(v, sn):
                    cell.fill = PatternFill('solid', start_color=FORMULA_FILL)
                    cell.font = new_font(LINK_FG)
                    stats['links'] += 1
                else:
                    cell.fill = PatternFill('solid', start_color=FORMULA_FILL)
                    cell.font = new_font(FORMULA_FG)
                    stats['formulas'] += 1
            elif isinstance(v, (int, float)):
                # Hardcoded numeric input
                cell.fill = PatternFill('solid', start_color=INPUT_FILL)
                # keep original font color for inputs
                cell.font = new_font(cell.font.color.rgb if (cell.font.color and isinstance(cell.font.color.rgb, str)) else None)
                stats['inputs'] += 1
            else:
                # text label — just rename the font, leave style alone
                cell.font = new_font(cell.font.color.rgb if (cell.font.color and isinstance(cell.font.color.rgb, str)) else None)

print("Coloring applied:", stats)

# ---------------------------------------------------------------------------
# 2) MASTER "Color Key" TAB  (insert at position 0)
# ---------------------------------------------------------------------------
thin = Side(style='thin', color='BFBFBF')
box  = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_bottom = Border(bottom=Side(style='medium', color=UH_GREEN))

if 'Color Key' in wb.sheetnames:
    del wb['Color Key']
key = wb.create_sheet('Color Key', 0)

key.sheet_view.showGridLines = False
key.sheet_view.zoomScale = 110
key.column_dimensions['A'].width = 2
key.column_dimensions['B'].width = 22
key.column_dimensions['C'].width = 26
key.column_dimensions['D'].width = 70
key.column_dimensions['E'].width = 2

# Title banner
key.merge_cells('B2:D2')
t = key['B2']
t.value = 'Color Key & Legend — Corporate Finance Master Workbook'
t.font = Font(name='Open Sans', size=20, bold=True, color=UH_WHITE)
t.fill = PatternFill('solid', start_color=UH_GREEN)
t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
key.row_dimensions[2].height = 38

key.merge_cells('B3:D3')
s = key['B3']
s.value = ('This workbook is a resource for BUS-313, BUS-314, and FIN-321. '
          'Tabs are grouped by chapter. Colors follow a consistent convention '
          'so you can tell at a glance which cells to edit and which to leave alone.')
s.font = Font(name='Open Sans', size=11, italic=True, color='595959')
s.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)
key.row_dimensions[3].height = 36

# Table header
hdrs = [('B',5,'Cell Type'),('C',5,'Color'),('D',5,'Meaning / When You Will See It')]
for col,row,text in hdrs:
    c = key[f'{col}{row}']
    c.value = text
    c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
    c.fill = PatternFill('solid', start_color=UH_GREEN)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = box

legend_rows = [
    ('Input',         INPUT_FILL,   None,        '1,000',
     'Hard-coded values — these are the assumptions you can change to explore scenarios. '
     'Every yellow cell is an input.'),
    ('Formula',       FORMULA_FILL, FORMULA_FG,  '=B5*(1+rate)',
     'Computed values. Blue text on gray fill. Do not edit — these update automatically.'),
    ('Array formula', FORMULA_FILL, ARRAY_FG,    '{=INDIRECT(...)}',
     'Dynamic lookup using named ranges (common on the Ratios tab). Purple text on gray fill.'),
    ('Linked cell',   FORMULA_FILL, LINK_FG,     "='Balance Sheet'!H12",
     'Formula that pulls a value from another worksheet. Green text on gray fill.'),
    ('Header / label',UH_GREEN,     UH_WHITE,    'CHAPTER 5',
     'Titles and section headings. UH green. Not calculated.'),
    ('Table header',  'D9D9D9',     '000000',    'Metric  |  Input',
     'Column headings on data tables. Dark gray.'),
]

for i, (label, fill, font_rgb, example, meaning) in enumerate(legend_rows):
    r = 6 + i
    key.row_dimensions[r].height = 30

    a = key.cell(row=r, column=2, value=label)
    a.font = Font(name='Open Sans', size=11, bold=True)
    a.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    a.border = box

    b = key.cell(row=r, column=3, value=example)
    b.fill = PatternFill('solid', start_color=fill)
    if font_rgb:
        b.font = Font(name='Open Sans', size=11, color=font_rgb, bold=(label=='Header / label'))
    else:
        b.font = Font(name='Open Sans', size=11)
    b.alignment = Alignment(horizontal='center', vertical='center')
    b.border = box

    c = key.cell(row=r, column=4, value=meaning)
    c.font = Font(name='Open Sans', size=11, color='333333')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    c.border = box

# Footer / guidance
key.merge_cells('B13:D13')
tips = key['B13']
tips.value = 'Tip for students'
tips.font = Font(name='Open Sans', size=12, bold=True, color=UH_GREEN)
tips.alignment = Alignment(horizontal='left', vertical='center', indent=1)
tips.border = thick_bottom

key.merge_cells('B14:D16')
tb = key['B14']
tb.value = ('If a cell is yellow, you may change it to test a what-if. '
            'Everything else is either a label or a calculation that depends on '
            'other cells — editing those will break the model. Use the chapter '
            'tabs at the bottom of the workbook to jump to a topic.')
tb.font = Font(name='Open Sans', size=11, color='333333')
tb.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)

key.freeze_panes = 'A5'

# Tab color — UH green
key.sheet_properties.tabColor = UH_GREEN

# ---------------------------------------------------------------------------
# 3) COMPACT LEGEND on each chapter section tab (columns G-J, rows 5+)
# ---------------------------------------------------------------------------
compact_rows = [
    ('Input',         INPUT_FILL,   None),
    ('Formula',       FORMULA_FILL, FORMULA_FG),
    ('Array formula', FORMULA_FILL, ARRAY_FG),
    ('Linked cell',   FORMULA_FILL, LINK_FG),
]

for sn in chapter_section_tabs:
    ws = wb[sn]
    # Header
    ws.merge_cells('G5:I5')
    h = ws['G5']
    h.value = 'Color key'
    h.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
    h.fill = PatternFill('solid', start_color=UH_GREEN)
    h.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    h.border = box

    for i,(label, fill, fg) in enumerate(compact_rows):
        r = 6 + i
        swatch = ws.cell(row=r, column=7, value='')
        swatch.fill = PatternFill('solid', start_color=fill)
        swatch.border = box
        nm = ws.cell(row=r, column=8, value=label)
        nm.font = Font(name='Open Sans', size=10, color=fg or '000000')
        nm.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        nm.border = box

    # Footer: point to full key
    ws.merge_cells(f'G{6+len(compact_rows)}:I{6+len(compact_rows)}')
    link_c = ws.cell(row=6+len(compact_rows), column=7, value='See the "Color Key" tab for full legend')
    link_c.hyperlink = "#'Color Key'!A1"
    link_c.font = Font(name='Open Sans', size=9, italic=True, color=UH_GREEN, underline='single')
    link_c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # Column widths for the key area
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 4

wb.save(SRC)
print("openpyxl save complete.")

# ---------------------------------------------------------------------------
# 4) POST-SAVE: remap localSheetId worksheet-only -> full index
#    AND patch xl/styles.xml to replace font name Calibri -> Open Sans globally
# ---------------------------------------------------------------------------
# We cannot re-load with openpyxl here because the saved localSheetId values
# point to worksheet-only positions; when openpyxl tries to attach those local
# names to sheets at the corresponding full-list positions, several land on
# chartsheets and the load fails.  Parse the XML directly instead.
tmp = SRC.with_suffix('.xlsx.tmp')
with zipfile.ZipFile(SRC, 'r') as zin:
    names = zin.namelist()
    buf = {n: zin.read(n) for n in names}

# Parse sheet list & relationships from the saved file
_rels_xml = buf['xl/_rels/workbook.xml.rels'].decode('utf-8')
_ct_xml   = buf['[Content_Types].xml'].decode('utf-8')
rel_target = {}
for r in re.finditer(r'<Relationship\b[^>]*?>', _rels_xml):
    tag = r.group(0)
    id_m = re.search(r'Id="([^"]+)"', tag)
    tgt_m = re.search(r'Target="([^"]+)"', tag)
    if id_m and tgt_m:
        rel_target[id_m.group(1)] = tgt_m.group(1)
ct_type = {}
for o in re.finditer(r'<Override\b[^>]*?>', _ct_xml):
    tag = o.group(0)
    pn = re.search(r'PartName="([^"]+)"', tag)
    ct = re.search(r'ContentType="([^"]+)"', tag)
    if pn and ct:
        ct_type[pn.group(1)] = ct.group(1)

# Iterate sheets in workbook.xml order and flag chartsheets
_wb_xml_raw = buf['xl/workbook.xml'].decode('utf-8')
sheet_flags = []
for m in re.finditer(r'<sheet\b[^>]*?>', _wb_xml_raw):
    tag = m.group(0)
    rm = re.search(r'r:id="([^"]+)"', tag)
    if not rm:
        sheet_flags.append(False)
        continue
    target = rel_target.get(rm.group(1), '')
    part = target if target.startswith('/') else ('/xl/' + target.lstrip('/'))
    ctype = ct_type.get(part, '')
    sheet_flags.append('chart' in ctype.lower())

new_ws_only_to_full = [i for i,ch in enumerate(sheet_flags) if not ch]
print(f"Sheets after save: {len(sheet_flags)} total, "
      f"{sum(sheet_flags)} charts, {len(new_ws_only_to_full)} worksheets")

# 4a) workbook.xml localSheetId remap (worksheet-only idx -> full idx)
wb_xml = _wb_xml_raw
def remap(m):
    ws_idx = int(m.group(1))
    if ws_idx < len(new_ws_only_to_full):
        return f'localSheetId="{new_ws_only_to_full[ws_idx]}"'
    return m.group(0)
wb_xml = re.sub(r'localSheetId="(\d+)"', remap, wb_xml)

# 4a-2) drop orphan global definedNames that point to non-existent external
# workbooks ([1]...) or raw #REF! -- openpyxl can promote stale local-scope
# names into the global namespace during save.
orphan_pattern = r'<definedName\b(?![^>]*localSheetId=)[^>]*>[^<]*(?:\[1\]|#REF!)[^<]*</definedName>'
orphans = re.findall(orphan_pattern, wb_xml)
wb_xml = re.sub(orphan_pattern, '', wb_xml)
print(f"Removed {len(orphans)} orphan global named-ranges from workbook.xml")
buf['xl/workbook.xml'] = wb_xml.encode('utf-8')

# 4b) styles.xml font name -> Open Sans (keep size/bold/color)
styles_xml = buf['xl/styles.xml'].decode('utf-8')
n_font = len(re.findall(r'<name\s+val="Calibri"\s*/>', styles_xml))
styles_xml = re.sub(r'<name\s+val="Calibri"\s*/>', '<name val="Open Sans"/>', styles_xml)
styles_xml = re.sub(r'<name\s+val="Arial"\s*/>', '<name val="Open Sans"/>', styles_xml)
styles_xml = re.sub(r'<name\s+val="Inconsolata"\s*/>', '<name val="Open Sans"/>', styles_xml)
styles_xml = re.sub(r'<name\s+val="Roboto"\s*/>', '<name val="Open Sans"/>', styles_xml)
buf['xl/styles.xml'] = styles_xml.encode('utf-8')
print(f"Font-name replacements in styles.xml: {n_font} Calibri -> Open Sans (plus Arial/Inconsolata/Roboto)")

with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
    for n in names:
        zout.writestr(n, buf[n])
shutil.move(str(tmp), str(SRC))
print("Post-save fixups complete.")
