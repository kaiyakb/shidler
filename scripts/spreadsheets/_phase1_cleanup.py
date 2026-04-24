"""Phase 1: safe cleanup of Corporate Finance Master Spreadsheets.xlsx
  1) Fix 21 broken named ranges (repoint 9, delete 13 orphans)
  2) Replace _xludf.FORMULATEXT with _xlfn.FORMULATEXT (fixes #NAME?)
  3) Wrap tail-row formulas in IFERROR to kill #DIV/0!
  4) Rebuild 9 chapter section tabs as proper dividers with clickable ToC

Post-save fixup: openpyxl writes `localSheetId` using a worksheet-only index,
but Excel reads it against the full sheet list (including chartsheets).  After
saving we remap each localSheetId to the full-list position so the file
round-trips correctly.
"""
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chartsheet import Chartsheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = Path(r'C:/GitHub/shidler/docs/spreadsheets/Corporate Finance Master Spreadsheets.xlsx')

UH_GREEN = '024731'
UH_WHITE = 'FFFFFF'
LIGHT_GREEN = 'E6EEEA'

wb = load_workbook(SRC, data_only=False)

# Record full sheet order & chartsheet positions BEFORE we touch anything.
full_order = list(wb.sheetnames)
is_chart = [isinstance(wb[n], Chartsheet) for n in full_order]
worksheet_only_to_full = [i for i, ch in enumerate(is_chart) if not ch]
# worksheet_only_to_full[k] = the full-list index of the k-th non-chart sheet
# e.g. worksheet_only_to_full[62] = 69 in this workbook

# ---------------------------------------------------------------------------
# 1) NAMED RANGES
# ---------------------------------------------------------------------------
repoint = {
    'beta':         "'Value Business (step 4)'!$E$11",
    'rf_rate':      "'Value Business (step 4)'!$E$12",
    'risk_premium': "'Value Business (step 4)'!$E$13",
    'tax':          "'Value Business (step 4)'!$E$14",
    'rtn_debt':     "'Value Business (step 4)'!$C$18",
    'rtn_equity':   "'Value Business (step 4)'!$C$19",
    'wacc':         "'Value Business (step 4)'!$C$20",
    'growth_rate':  "'Value Business (step 4)'!$I$24",
}
delete_names = [
    'age','age_retirement','birthday','date_retirement','days_to_retirement',
    'retirement_objective','save_annual','savings_current','spending_retirement',
    'years_in_retirement','years_to_retirement',
    'PT_income','PT_tax',
]
for n, target in repoint.items():
    if n in wb.defined_names:
        del wb.defined_names[n]
    wb.defined_names[n] = DefinedName(name=n, attr_text=target)
for n in delete_names:
    if n in wb.defined_names:
        del wb.defined_names[n]
print(f"Named ranges: {len(list(wb.defined_names))} global (was 184)")

# ---------------------------------------------------------------------------
# 2) _xludf.FORMULATEXT -> _xlfn.FORMULATEXT
# ---------------------------------------------------------------------------
xludf = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    if isinstance(ws, Chartsheet):
        continue
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and '_xludf.' in v:
                cell.value = v.replace('_xludf.', '_xlfn.')
                xludf += 1
print(f"_xludf replacements: {xludf}")

# ---------------------------------------------------------------------------
# 3) #DIV/0! tail-row IFERROR wraps
# ---------------------------------------------------------------------------
div_fixes = [
    ('Risk Correl ZM + UAL', 'J503'),
    ('Risk Correl ZM + UAL', 'M503'),
    ('Risk Correl ZM + UAL', 'J504'),
    ('Risk Correl ZM + UAL', 'M504'),
    ('GBTC', 'D756'),('GLD', 'D756'),('IEF', 'D756'),('SPY', 'D756'),
    ('Asset Classes', 'N756'),('Asset Classes', 'O756'),
    ('Asset Classes wBTC', 'N756'),('Asset Classes wBTC', 'O756'),
]
for sn, coord in div_fixes:
    c = wb[sn][coord]
    f = c.value
    if isinstance(f, str) and f.startswith('=') and 'IFERROR' not in f.upper():
        c.value = f'=IFERROR({f[1:]},"")'
print(f"IFERROR-wrapped cells: {len(div_fixes)}")

# ---------------------------------------------------------------------------
# 4) Chapter section tabs -> dividers with ToC
# ---------------------------------------------------------------------------
def tabs_between(start, end):
    i0 = full_order.index(start) + 1
    return full_order[i0:] if end is None else full_order[i0:full_order.index(end)]

chapter_map = [
    ('Chapter 3 + 4', 'Chapter 3 & 4 — Accounting and Finance; Measuring Corporate Performance', 'Chapter 5'),
    ('Chapter 5', 'Chapter 5 — The Time Value of Money', 'Chapter 6'),
    ('Chapter 6', 'Chapter 6 — Valuing Bonds', 'Chapter 7 Valuing Stocks'),
    ('Chapter 7 Valuing Stocks', 'Chapter 7 — Valuing Stocks', 'Chapter 8 NPV'),
    ('Chapter 8 NPV', 'Chapter 8 — Net Present Value and Other Investment Criteria', 'Chapter 11 Intro to Risk'),
    ('Chapter 11 Intro to Risk', 'Chapter 11 — Introduction to Risk, Return, and the Opportunity Cost of Capital', 'Chapter 12 Risk, Return, Budget'),
    ('Chapter 12 Risk, Return, Budget', 'Chapter 12 — Risk, Return, and Capital Budgeting', 'Chapter 13 WACC'),
    ('Chapter 13 WACC', 'Chapter 13 — The Weighted-Average Cost of Capital and Company Valuation', 'Chapter 23 Options'),
    ('Chapter 23 Options', 'Chapter 23 — Options (and Derivatives)', None),
]
thin = Side(style='thin', color='BFBFBF')
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def clear_sheet(ws):
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

for section_tab, title, nxt in chapter_map:
    ws = wb[section_tab]
    members = tabs_between(section_tab, nxt)
    clear_sheet(ws)

    ws.merge_cells('B2:E2')
    t = ws['B2']
    t.value = title
    t.font = Font(name='Open Sans', size=18, bold=True, color=UH_WHITE)
    t.fill = PatternFill('solid', start_color=UH_GREEN)
    t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 34

    ws.merge_cells('B3:E3')
    s = ws['B3']
    s.value = 'Contents — click a tab name to jump to that sheet'
    s.font = Font(name='Open Sans', size=11, italic=True, color='595959')
    s.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[3].height = 20

    for i, h in enumerate(['#', 'Tab Name', 'Type'], start=2):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
        c.fill = PatternFill('solid', start_color=UH_GREEN)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = box

    for idx, tab in enumerate(members, start=1):
        r = 5 + idx
        kind = 'Chart' if isinstance(wb[tab], Chartsheet) else 'Worksheet'
        n = ws.cell(row=r, column=2, value=idx)
        n.font = Font(name='Open Sans', size=11, color='595959')
        n.alignment = Alignment(horizontal='center', vertical='center')
        n.border = box
        link = ws.cell(row=r, column=3, value=tab)
        safe = tab.replace("'", "''")
        link.hyperlink = f"#'{safe}'!A1"
        link.font = Font(name='Open Sans', size=11, color='024731', underline='single')
        link.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        link.border = box
        tp = ws.cell(row=r, column=4, value=kind)
        tp.font = Font(name='Open Sans', size=11, color='595959')
        tp.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        tp.border = box
        if idx % 2 == 0:
            for col in (2, 3, 4):
                ws.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 2
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.freeze_panes = 'B6'

    print(f"  rebuilt {section_tab}: {len(members)} tabs")

wb.save(SRC)
print("openpyxl save complete.")

# ---------------------------------------------------------------------------
# POST-SAVE FIXUP: remap localSheetId from worksheet-only index -> full index
# ---------------------------------------------------------------------------
tmp = SRC.with_suffix('.xlsx.tmp')

with zipfile.ZipFile(SRC, 'r') as zin:
    names = zin.namelist()
    buffers = {n: zin.read(n) for n in names}

wb_xml = buffers['xl/workbook.xml'].decode('utf-8')

def remap(m):
    ws_idx = int(m.group(1))
    # translate worksheet-only idx -> full idx
    full_idx = worksheet_only_to_full[ws_idx] if ws_idx < len(worksheet_only_to_full) else ws_idx
    return f'localSheetId="{full_idx}"'

new_wb_xml = re.sub(r'localSheetId="(\d+)"', remap, wb_xml)
buffers['xl/workbook.xml'] = new_wb_xml.encode('utf-8')

with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
    for n in names:
        zout.writestr(n, buffers[n])
shutil.move(str(tmp), str(SRC))
print("Post-save localSheetId remap complete.")
