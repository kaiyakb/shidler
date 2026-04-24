"""Phase 3: structural / pedagogical cleanup
  (a) Remove 11 stray 'Chapter N: ...' labels on content tabs (Ratios, the
      financial statements, and the two class-example tabs)
  (b) Add a Welcome / How-to-use tab as sheet #1 (Color Key becomes #2)
  (c) Color each tab by chapter grouping for visual orientation
  (d) Hide the raw ETF price-data tabs (students don't need to see them)
  (e) Preserve everything else; renames are deferred to avoid breaking
      formula references
"""
import re, shutil, zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = Path(r'C:/GitHub/shidler/docs/spreadsheets/Corporate Finance Master Spreadsheets.xlsx')

UH_GREEN    = '024731'
UH_WHITE    = 'FFFFFF'
LIGHT_GREEN = 'E6EEEA'

wb = load_workbook(SRC, data_only=False)

# ---------------------------------------------------------------------------
# (a) STRAY CHAPTER LABELS
# ---------------------------------------------------------------------------
strays = [
    ('Market value (class example)', 'L2'),
    ('Market value (class example)', 'L3'),
    ('Free Cash Flow (class example)', 'L2'),
    ('Free Cash Flow (class example)', 'L3'),
    ('Balance Sheet', 'L3'),
    ('Income Statement', 'L2'),
    ('Income Statement', 'L3'),
    ('Cash Flow Statement', 'L2'),
    ('Cash Flow Statement', 'L3'),
    ('Ratios', 'L2'),
    ('Ratios', 'L3'),
]
for sn, coord in strays:
    wb[sn][coord].value = None
print(f"Removed stray labels: {len(strays)}")

# ---------------------------------------------------------------------------
# (b) WELCOME TAB
# ---------------------------------------------------------------------------
if 'Welcome' in wb.sheetnames:
    del wb['Welcome']
w = wb.create_sheet('Welcome', 0)

w.sheet_view.showGridLines = False
w.sheet_view.zoomScale = 110
w.column_dimensions['A'].width = 2
w.column_dimensions['B'].width = 8
w.column_dimensions['C'].width = 40
w.column_dimensions['D'].width = 60
w.column_dimensions['E'].width = 2

# Title banner
w.merge_cells('B2:D2')
t = w['B2']
t.value = 'Corporate Finance Master Workbook'
t.font = Font(name='Open Sans', size=22, bold=True, color=UH_WHITE)
t.fill = PatternFill('solid', start_color=UH_GREEN)
t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
w.row_dimensions[2].height = 42

w.merge_cells('B3:D3')
s = w['B3']
s.value = 'UH Mānoa · Shidler College of Business · BUS 313 / BUS 314 / FIN 321 / BUS 620 / BUS 629'
s.font = Font(name='Open Sans', size=11, italic=True, color='595959')
s.alignment = Alignment(horizontal='left', vertical='center', indent=1)
w.row_dimensions[3].height = 20

# Intro paragraph
w.merge_cells('B5:D7')
intro = w['B5']
intro.value = ('This workbook is a reference library of worked examples for corporate '
               'finance. Each chapter matches a chapter of Brealey, Myers, Marcus — '
               'Fundamentals of Corporate Finance. Click a chapter below to jump to its '
               'contents page, which lists every example tab for that chapter. The '
               'Color Key tab explains how cells are colored (yellow = change these; '
               'gray/blue = calculated — leave alone).')
intro.font = Font(name='Open Sans', size=11, color='333333')
intro.alignment = Alignment(horizontal='left', vertical='top', indent=1, wrap_text=True)

# Navigation table
thin = Side(style='thin', color='BFBFBF')
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

w.merge_cells('B9:D9')
hx = w['B9']
hx.value = 'Chapter Navigation'
hx.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
hx.alignment = Alignment(horizontal='left', vertical='center', indent=1)
hx.border = Border(bottom=Side(style='medium', color=UH_GREEN))

# Headers
for col, text in zip(['B','C','D'], ['Chapter', 'Topic', 'What you will find']):
    c = w[f'{col}10']
    c.value = text
    c.font = Font(name='Open Sans', size=11, bold=True, color=UH_WHITE)
    c.fill = PatternFill('solid', start_color=UH_GREEN)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = box

chapter_nav = [
    ('Chapter 3 + 4',               '3 & 4 — Accounting & Performance',             'Balance sheet, income statement, cash flow, and 25+ performance ratios tied together via named ranges.'),
    ('Chapter 5',                   '5 — Time Value of Money',                       '15 worked examples: FV, PV, perpetuities, annuities, mortgages, retirement, inflation.'),
    ('Chapter 6',                   '6 — Valuing Bonds',                             'Annual vs semi-annual, YTM, rate of return, comparisons across coupons and maturities.'),
    ('Chapter 7 Valuing Stocks',    '7 — Valuing Stocks',                            'Intrinsic value, DDM, constant-growth, non-constant-growth, PVGO.'),
    ('Chapter 8 NPV',               '8 — NPV & Investment Criteria',                 'Office building case, IRR, project comparison.'),
    ('Chapter 11 Intro to Risk',    '11 — Intro to Risk & Return',                    'S&P histogram, ETF price data, correlations, asset-class comparisons.'),
    ('Chapter 12 Risk, Return, Budget', '12 — Risk, Return & Capital Budgeting',      'Beta (TSLA vs SPX), project-risk case study.'),
    ('Chapter 13 WACC',             '13 — WACC & Company Valuation',                 'WACC with 2 and 3 securities; four-step DCF valuation.'),
    ('Chapter 23 Options',          '23 — Options & Derivatives',                    'Payoff diagrams for long/short calls and puts; currency hedge.'),
]
for i, (tab_name, topic, desc) in enumerate(chapter_nav):
    r = 11 + i
    w.row_dimensions[r].height = 30

    b = w.cell(row=r, column=2, value=tab_name)
    safe = tab_name.replace("'", "''")
    b.hyperlink = f"#'{safe}'!A1"
    b.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN, underline='single')
    b.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    b.border = box

    cc = w.cell(row=r, column=3, value=topic)
    cc.font = Font(name='Open Sans', size=11)
    cc.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    cc.border = box

    dd = w.cell(row=r, column=4, value=desc)
    dd.font = Font(name='Open Sans', size=10, color='595959')
    dd.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    dd.border = box

    if i % 2 == 0:
        for col in (2, 3, 4):
            w.cell(row=r, column=col).fill = PatternFill('solid', start_color=LIGHT_GREEN)

# "Getting started" callout
gs_row = 11 + len(chapter_nav) + 1
w.merge_cells(f'B{gs_row}:D{gs_row}')
gs = w[f'B{gs_row}']
gs.value = 'Getting started'
gs.font = Font(name='Open Sans', size=13, bold=True, color=UH_GREEN)
gs.alignment = Alignment(horizontal='left', vertical='center', indent=1)
gs.border = Border(bottom=Side(style='medium', color=UH_GREEN))

tips = [
    ('1.', 'Open the Color Key tab', 'The color legend explains which cells are safe to edit.'),
    ('2.', 'Pick a chapter tab from the list above', 'Each chapter page is a jump table to that chapter\'s examples.'),
    ('3.', 'Change a yellow input cell', 'Watch how every gray/blue formula cell updates.'),
    ('4.', 'Never edit a gray (formula) cell', 'If you accidentally do, press Ctrl+Z to undo.'),
]
for i,(num, title, body) in enumerate(tips):
    r = gs_row + 1 + i
    w.row_dimensions[r].height = 22
    a = w.cell(row=r, column=2, value=num)
    a.font = Font(name='Open Sans', size=11, bold=True, color=UH_GREEN)
    a.alignment = Alignment(horizontal='center', vertical='center')
    t1 = w.cell(row=r, column=3, value=title)
    t1.font = Font(name='Open Sans', size=11, bold=True)
    t1.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    t2 = w.cell(row=r, column=4, value=body)
    t2.font = Font(name='Open Sans', size=10, color='595959')
    t2.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)

# Footer: link to Color Key
foot_row = gs_row + 1 + len(tips) + 2
w.merge_cells(f'B{foot_row}:D{foot_row}')
foot = w[f'B{foot_row}']
foot.value = '→ Open the Color Key tab'
foot.hyperlink = "#'Color Key'!A1"
foot.font = Font(name='Open Sans', size=11, italic=True, color=UH_GREEN, underline='single')
foot.alignment = Alignment(horizontal='left', vertical='center', indent=1)

w.sheet_properties.tabColor = UH_GREEN
w.freeze_panes = 'A5'

print("Welcome tab created at position 0")

# ---------------------------------------------------------------------------
# (c) TAB COLORS BY CHAPTER
# ---------------------------------------------------------------------------
# Assign a color to each chapter's tab group.  Welcome + Color Key keep UH green.
chapter_colors = {
    'ch_3_4':  '024731',  # UH green
    'ch_5':    '1F4E79',  # navy blue — TVM
    'ch_6':    '663398',  # purple — bonds
    'ch_7':    '117A8B',  # teal — stocks
    'ch_8':    'A0522D',  # sienna — NPV
    'ch_11':   '842029',  # deep red — risk
    'ch_12':   '831843',  # magenta — risk/budget
    'ch_13':   '8B6F00',  # dark gold — WACC
    'ch_23':   '3D3D3D',  # slate — options
}

# Map each tab to its chapter bucket based on its position relative to section tabs
section_tabs = ['Chapter 3 + 4','Chapter 5','Chapter 6','Chapter 7 Valuing Stocks',
                'Chapter 8 NPV','Chapter 11 Intro to Risk','Chapter 12 Risk, Return, Budget',
                'Chapter 13 WACC','Chapter 23 Options']
section_keys = ['ch_3_4','ch_5','ch_6','ch_7','ch_8','ch_11','ch_12','ch_13','ch_23']

order = wb.sheetnames
section_indices = [order.index(s) for s in section_tabs]

def chapter_of(index):
    if index < section_indices[0]:
        return None  # before first chapter section
    for i in range(len(section_indices)):
        start = section_indices[i]
        end = section_indices[i+1] if i+1 < len(section_indices) else len(order)
        if start <= index < end:
            return section_keys[i]
    return None

colored = 0
for i, name in enumerate(order):
    if name in ('Welcome', 'Color Key'):
        continue
    sheet = wb[name]
    # chartsheets don't expose sheet_properties in openpyxl — skip (their tab
    # color will default to unset and they are grouped by position anyway)
    if isinstance(sheet, Chartsheet):
        continue
    bucket = chapter_of(i)
    if bucket:
        sheet.sheet_properties.tabColor = chapter_colors[bucket]
        colored += 1
print(f"Colored tabs: {colored}")

# ---------------------------------------------------------------------------
# (d) HIDE RAW DATA TABS
# ---------------------------------------------------------------------------
hide = ['GBTC','GLD','IEF','SPY','EWJ','EZU','EWA','correl','data']
hidden_count = 0
for sn in hide:
    if sn in wb.sheetnames:
        wb[sn].sheet_state = 'hidden'
        hidden_count += 1
print(f"Hidden data tabs: {hidden_count}")

wb.save(SRC)
print("openpyxl save complete.")

# ---------------------------------------------------------------------------
# POST-SAVE FIXUP (same pattern as Phase 2)
# ---------------------------------------------------------------------------
tmp = SRC.with_suffix('.xlsx.tmp')
with zipfile.ZipFile(SRC, 'r') as zin:
    names = zin.namelist()
    buf = {n: zin.read(n) for n in names}

_rels_xml = buf['xl/_rels/workbook.xml.rels'].decode('utf-8')
_ct_xml   = buf['[Content_Types].xml'].decode('utf-8')
rel_target = {}
for r in re.finditer(r'<Relationship\b[^>]*?>', _rels_xml):
    tag = r.group(0)
    im = re.search(r'Id="([^"]+)"', tag); tm = re.search(r'Target="([^"]+)"', tag)
    if im and tm: rel_target[im.group(1)] = tm.group(1)
ct_type = {}
for o in re.finditer(r'<Override\b[^>]*?>', _ct_xml):
    tag = o.group(0)
    pm = re.search(r'PartName="([^"]+)"', tag); cm = re.search(r'ContentType="([^"]+)"', tag)
    if pm and cm: ct_type[pm.group(1)] = cm.group(1)

_wb_xml = buf['xl/workbook.xml'].decode('utf-8')
sheet_flags = []
for m in re.finditer(r'<sheet\b[^>]*?>', _wb_xml):
    tag = m.group(0)
    rm = re.search(r'r:id="([^"]+)"', tag)
    if not rm:
        sheet_flags.append(False); continue
    tgt = rel_target.get(rm.group(1), '')
    part = tgt if tgt.startswith('/') else ('/xl/' + tgt.lstrip('/'))
    sheet_flags.append('chart' in ct_type.get(part, '').lower())
new_ws_only_to_full = [i for i,ch in enumerate(sheet_flags) if not ch]
print(f"Sheets after save: {len(sheet_flags)} total, {sum(sheet_flags)} charts, "
      f"{len(new_ws_only_to_full)} worksheets")

# remap localSheetId
def remap(m):
    idx = int(m.group(1))
    return f'localSheetId="{new_ws_only_to_full[idx]}"' if idx < len(new_ws_only_to_full) else m.group(0)
_wb_xml = re.sub(r'localSheetId="(\d+)"', remap, _wb_xml)

# drop any newly-promoted orphan globals
op = r'<definedName\b(?![^>]*localSheetId=)[^>]*>[^<]*(?:\[1\]|#REF!)[^<]*</definedName>'
orphans = re.findall(op, _wb_xml)
_wb_xml = re.sub(op, '', _wb_xml)
print(f"Removed {len(orphans)} orphan global named-ranges")

buf['xl/workbook.xml'] = _wb_xml.encode('utf-8')

# re-apply font name normalization (Phase 2 already did it, but new cells
# created by this script come in as Calibri)
styles_xml = buf['xl/styles.xml'].decode('utf-8')
for old in ('Calibri','Arial','Inconsolata','Roboto'):
    styles_xml = re.sub(rf'<name\s+val="{old}"\s*/>', '<name val="Open Sans"/>', styles_xml)
buf['xl/styles.xml'] = styles_xml.encode('utf-8')

with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
    for n in names:
        zout.writestr(n, buf[n])
shutil.move(str(tmp), str(SRC))
print("Post-save fixups complete.")
